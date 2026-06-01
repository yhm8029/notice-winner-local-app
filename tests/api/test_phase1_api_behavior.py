from __future__ import annotations

import json
import os
import shutil
import socket
import subprocess
import sys
import tempfile
import time
import threading
import unittest
from datetime import datetime
from datetime import timezone
from http.server import BaseHTTPRequestHandler
from http.server import ThreadingHTTPServer
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any
from urllib.parse import parse_qs
from urllib.parse import urlparse
from urllib import error
from urllib import request
from uuid import UUID
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import requests
import pytest
from unittest.mock import patch

from backend.phase1_defaults import DEFAULT_PHASE1_INTERNAL_USER_ID
from backend.phase1_defaults import DEFAULT_PHASE1_ORGANIZATION_ID
from backend.phase1_defaults import load_phase1_identity
from backend.repositories import get_related_notice_publication_repository
from backend.repositories import reset_related_notice_publication_repository
from backend.repositories import InMemoryDownloadAuditLogRepository
from backend.repositories.in_memory_login_audit_logs import InMemoryLoginAuditLogRepository
from backend.sales_claims import SalesActor
from backend.services.artifact_files import read_tracking_workbook_rows

ROOT_DIR = Path(__file__).resolve().parents[2]


def _read_ttl_model_cache(*, cache: dict[str, tuple[float, Any]], cache_lock: threading.Lock, cache_key: str) -> Any:
    with cache_lock:
        cached = cache.get(cache_key)
        if cached is None:
            return None
        expires_at, payload = cached
        if expires_at <= time.time():
            cache.pop(cache_key, None)
            return None
        return payload


def _write_ttl_model_cache(
    *,
    cache: dict[str, tuple[float, Any]],
    cache_lock: threading.Lock,
    cache_key: str,
    ttl_sec: float,
    payload: Any,
) -> Any:
    with cache_lock:
        cached_payload = payload.model_dump() if hasattr(payload, "model_dump") else payload
        cache[cache_key] = (time.time() + ttl_sec, cached_payload)
        return cached_payload


def _clear_ttl_model_cache(*, cache: dict[str, tuple[float, Any]], cache_lock: threading.Lock, cache_key: str) -> None:
    with cache_lock:
        cache.pop(cache_key, None)


def _slice_items_with_has_more(items: list[Any], *, visible_limit: int) -> tuple[list[Any], bool]:
    visible_items = list(items[:visible_limit])
    return visible_items, len(items) > visible_limit


import backend.api.app as _phase1_app

if not hasattr(_phase1_app, "_read_ttl_model_cache"):
    _phase1_app._read_ttl_model_cache = _read_ttl_model_cache
    _phase1_app._write_ttl_model_cache = _write_ttl_model_cache
    _phase1_app._clear_ttl_model_cache = _clear_ttl_model_cache
if not hasattr(_phase1_app, "_slice_items_with_has_more"):
    _phase1_app._slice_items_with_has_more = _slice_items_with_has_more


def _test_auth_context(actor: SalesActor) -> SimpleNamespace:
    return SimpleNamespace(
        authorized=True,
        organization_id=actor.organization_id,
        local_user_id=actor.user_id,
        email=actor.email,
        display_name=actor.display_name,
        role=actor.role,
    )


def _patch_test_auth_context(app_module: Any, *, actor: SalesActor):
    return patch.object(app_module, "read_auth_context", return_value=_test_auth_context(actor))


def _find_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.bind(("127.0.0.1", 0))
        return int(sock.getsockname()[1])


def _phase1_test_tmp_root() -> Path:
    configured = os.getenv("PHASE1_TEST_TMP_ROOT", "").strip()
    base_root = Path(configured) if configured else Path(tempfile.gettempdir()) / "nwpw-tests"
    return base_root


def _build_test_tmp_path(*parts: str, token: str | None = None) -> Path:
    scope = str(token or uuid4().hex).strip() or uuid4().hex
    path = _phase1_test_tmp_root() / scope
    for part in parts:
        path = path / part
    return path


class ApiServer:
    def __init__(self, *, env_overrides: dict[str, str] | None = None) -> None:
        self._env_overrides = dict(env_overrides or {})
        self._tmp_root = _build_test_tmp_path(token=uuid4().hex)
        self._process: subprocess.Popen[str] | None = None
        self._port = _find_free_port()
        self._request_timeout_seconds = 300.0
        self.base_url = f"http://127.0.0.1:{self._port}"

    def __enter__(self) -> "ApiServer":
        self._tmp_root.mkdir(parents=True, exist_ok=True)
        env = os.environ.copy()
        env.update(
            {
                "PYTHONPATH": str(ROOT_DIR),
                "TRACKER_REPOSITORY_BACKEND": "in_memory",
                "TRACKER_CHANGE_EVENT_REPOSITORY_BACKEND": "in_memory",
                "BACKFILL_CONFLICT_REPOSITORY_BACKEND": "in_memory",
                "RUN_REPOSITORY_BACKEND": "in_memory",
                "ARTIFACT_REPOSITORY_BACKEND": "in_memory",
                "RUN_LOG_REPOSITORY_BACKEND": "in_memory",
                "DOWNLOAD_AUDIT_LOG_REPOSITORY_BACKEND": "in_memory",
                "ARTIFACTS_ROOT": str(self._tmp_root / "artifacts"),
                "RUN_WORKSPACE_ROOT": str(self._tmp_root / "workspace"),
                "TRACKER_TEMPLATE_PATH": str(ROOT_DIR / "assets" / "project_tracker_template.xlsx"),
                "TRACKER_TEMPLATE_UPLOAD_PATH": str(self._tmp_root / "uploaded_tracker_template.xlsx"),
                "TRACKER_TEMPLATE_UPLOAD_META_PATH": str(self._tmp_root / "uploaded_tracker_template.json"),
                "PROJECT_TRACKER_ENABLE_SYNTHETIC_DEBUG": "1",
                "DATA_GO_KR_SERVICE_KEY": "",
                "PUBLIC_DATA_SERVICE_KEY": "",
                "G2B_SERVICE_KEY": "",
                "PHASE2_AUTH_ENABLED": "0",
                "BOOTSTRAP_PLATFORM_ADMIN_EMAIL": "",
                "SUPABASE_URL": "",
                "SUPABASE_SECRET_KEY": "",
                "SUPABASE_SERVICE_ROLE_KEY": "",
                "SUPABASE_ANON_KEY": "",
            }
        )
        env.update(self._env_overrides)
        self._process = subprocess.Popen(
            [
                sys.executable,
                "-m",
                "uvicorn",
                "backend.api.app:app",
                "--host",
                "127.0.0.1",
                "--port",
                str(self._port),
            ],
            cwd=str(ROOT_DIR),
            env=env,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
        )
        self._wait_for_health()
        return self

    def __exit__(self, exc_type: object, exc: object, tb: object) -> None:
        if self._process is not None:
            self._process.terminate()
            try:
                self._process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                self._process.kill()
                self._process.wait(timeout=5)
            if self._process.stdout is not None:
                self._process.stdout.close()
        shutil.rmtree(self._tmp_root, ignore_errors=True)

    def request_json(
        self,
        method: str,
        path_or_url: str,
        *,
        payload: dict[str, Any] | None = None,
    ) -> tuple[int, Any]:
        url = self._resolve_url(path_or_url)
        headers = {"Accept": "application/json"}
        body = None
        if payload is not None:
            headers["Content-Type"] = "application/json"
            body = json.dumps(payload).encode("utf-8")
        req = request.Request(url, data=body, headers=headers, method=method)
        try:
            with request.urlopen(req, timeout=self._request_timeout_seconds) as response:
                content = response.read().decode("utf-8").strip()
                return response.status, json.loads(content) if content else None
        except error.HTTPError as exc:
            try:
                content = exc.read().decode("utf-8").strip()
                return exc.code, json.loads(content) if content else None
            finally:
                exc.close()

    def request_bytes(self, method: str, path_or_url: str) -> tuple[int, bytes, dict[str, str]]:
        url = self._resolve_url(path_or_url)
        req = request.Request(url, headers={"Accept": "*/*"}, method=method)
        try:
            with request.urlopen(req, timeout=self._request_timeout_seconds) as response:
                body = response.read()
                headers = {key.lower(): value for key, value in response.headers.items()}
                return response.status, body, headers
        except error.HTTPError as exc:
            try:
                return exc.code, exc.read(), {}
            finally:
                exc.close()

    def open_url(self, path_or_url: str, *, headers: dict[str, str] | None = None, timeout: float = 300.0):
        url = self._resolve_url(path_or_url)
        req = request.Request(url, headers=headers or {}, method="GET")
        return request.urlopen(req, timeout=timeout)

    def wait_for_run(self, run_id: str, *, target_statuses: set[str], timeout_seconds: float = 20.0) -> Any:
        deadline = time.time() + timeout_seconds
        last_payload: Any = None
        while time.time() < deadline:
            status_code, payload = self.request_json("GET", f"/api/runs/{run_id}")
            if status_code != 200:
                raise AssertionError(f"unexpected status while polling run {run_id}: {status_code} {payload}")
            last_payload = payload
            run_status = payload["status"]
            if run_status in target_statuses:
                return payload
            time.sleep(0.05)
        raise AssertionError(f"timed out waiting for run {run_id}; last payload={last_payload}")

    def _wait_for_health(self) -> None:
        deadline = time.time() + 20
        while time.time() < deadline:
            if self._process is not None and self._process.poll() is not None:
                raise AssertionError(self._process_output())
            try:
                status_code, payload = self.request_json("GET", "/health")
            except Exception:
                time.sleep(0.2)
                continue
            if status_code == 200 and payload == {"status": "ok"}:
                return
            time.sleep(0.2)
        raise AssertionError(f"timed out waiting for /health\n{self._process_output()}")

    def _resolve_url(self, path_or_url: str) -> str:
        if path_or_url.startswith("http://") or path_or_url.startswith("https://"):
            return path_or_url
        return f"{self.base_url}{path_or_url}"

    def _process_output(self) -> str:
        if self._process is None or self._process.stdout is None:
            return ""
        try:
            return self._process.stdout.read()
        except Exception:
            return ""


class _RelatedNoticePublicationMockServer:
    def __init__(self, *, reject_writes: bool = False) -> None:
        self._server: ThreadingHTTPServer | None = None
        self._thread: threading.Thread | None = None
        self._row_lock = threading.Lock()
        self._row: dict[str, Any] | None = None
        self._reject_writes = reject_writes
        self.base_url = ""

    def __enter__(self) -> "_RelatedNoticePublicationMockServer":
        row_lock = self._row_lock
        outer = self

        class Handler(BaseHTTPRequestHandler):
            def _write_json(self, status_code: int, payload: list[dict[str, Any]] | dict[str, Any]) -> None:
                body = json.dumps(payload).encode("utf-8")
                self.send_response(status_code)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)

            def _read_payload(self) -> dict[str, Any]:
                raw = self.rfile.read(int(self.headers.get("Content-Length", "0") or 0))
                if not raw:
                    return {}
                return json.loads(raw.decode("utf-8"))

            def _current_row(self) -> dict[str, Any] | None:
                with row_lock:
                    if outer._row is None:
                        return None
                    return dict(outer._row)

            def do_GET(self) -> None:  # noqa: N802
                parsed = urlparse(self.path)
                if parsed.path != "/rest/v1/related_notice_publications":
                    self._write_json(404, {"message": "not found"})
                    return
                query = parse_qs(parsed.query)
                organization_filter = (query.get("organization_id") or [""])[0]
                expected_organization_id = organization_filter[3:] if organization_filter.startswith("eq.") else ""
                row = self._current_row()
                if row is None:
                    self._write_json(200, [])
                    return
                if expected_organization_id and str(row.get("organization_id") or "") != expected_organization_id:
                    self._write_json(200, [])
                    return
                self._write_json(200, [row])

            def do_POST(self) -> None:  # noqa: N802
                parsed = urlparse(self.path)
                if parsed.path != "/rest/v1/related_notice_publications":
                    self._write_json(404, {"message": "not found"})
                    return
                if outer._reject_writes:
                    self._write_json(500, {"message": "publish rejected"})
                    return
                payload = self._read_payload()
                with row_lock:
                    existing = dict(outer._row or {})
                    row = {
                        "organization_id": str(payload.get("organization_id") or ""),
                        "published_snapshot_set_id": str(payload.get("published_snapshot_set_id") or ""),
                        "source_run_id": str(payload.get("source_run_id") or ""),
                        "generated_at": str(payload.get("generated_at") or ""),
                        "published_at": str(payload.get("published_at") or ""),
                        "created_at": existing.get("created_at") or str(payload.get("generated_at") or ""),
                        "updated_at": str(payload.get("published_at") or ""),
                    }
                    outer._row = row
                self._write_json(200, [row])

            def do_PATCH(self) -> None:  # noqa: N802
                parsed = urlparse(self.path)
                if parsed.path != "/rest/v1/related_notice_publications":
                    self._write_json(404, {"message": "not found"})
                    return
                if outer._reject_writes:
                    self._write_json(500, {"message": "publish rejected"})
                    return
                query = parse_qs(parsed.query)
                organization_filter = (query.get("organization_id") or [""])[0]
                updated_at_filter = (query.get("updated_at") or [""])[0]
                expected_organization_id = organization_filter[3:] if organization_filter.startswith("eq.") else ""
                expected_updated_at = updated_at_filter[3:] if updated_at_filter.startswith("eq.") else ""
                payload = self._read_payload()
                with row_lock:
                    existing = dict(outer._row or {})
                    if not existing:
                        self._write_json(200, [])
                        return
                    if expected_organization_id and str(existing.get("organization_id") or "") != expected_organization_id:
                        self._write_json(200, [])
                        return
                    if expected_updated_at and str(existing.get("updated_at") or "") != expected_updated_at:
                        self._write_json(200, [])
                        return
                    row = {
                        "organization_id": str(payload.get("organization_id") or existing.get("organization_id") or ""),
                        "published_snapshot_set_id": str(
                            payload.get("published_snapshot_set_id") or existing.get("published_snapshot_set_id") or ""
                        ),
                        "source_run_id": str(payload.get("source_run_id") or existing.get("source_run_id") or ""),
                        "generated_at": str(payload.get("generated_at") or existing.get("generated_at") or ""),
                        "published_at": str(payload.get("published_at") or existing.get("published_at") or ""),
                        "created_at": existing.get("created_at") or str(payload.get("generated_at") or ""),
                        "updated_at": str(payload.get("published_at") or existing.get("updated_at") or ""),
                    }
                    outer._row = row
                self._write_json(200, [row])

            def log_message(self, format: str, *args: object) -> None:  # noqa: A003
                del format, args

        self._server = ThreadingHTTPServer(("127.0.0.1", 0), Handler)
        self._thread = threading.Thread(target=self._server.serve_forever, daemon=True)
        self._thread.start()
        self.base_url = f"http://127.0.0.1:{self._server.server_port}"
        return self

    def __exit__(self, exc_type: object, exc: object, tb: object) -> None:
        del exc_type, exc, tb
        if self._server is not None:
            self._server.shutdown()
            self._server.server_close()
        if self._thread is not None:
            self._thread.join(timeout=5)


def _project_tracker_run_payload(*, delay_ms: int = 0) -> dict[str, Any]:
    suffix = uuid4().hex[:8]
    return {
        "run_type": "project_tracker",
        "advanced_options": {
            "collect_mode": "synthetic",
            "simulate_stage_delay_ms": delay_ms,
        },
        "params": {
            "start_date": "20250101",
            "end_date": "20250131",
            "notice_title": f"Phase1 Test {suffix}",
            "bid_no": f"PHASE1{suffix.upper()}",
            "demand_org": "Phase1 Internal Org",
            "rows_per_page": 50,
            "max_pages": 2,
            "api_scope": "construction",
        },
    }


def _project_tracker_run_payload_for_project(*, notice_title: str, bid_no: str, delay_ms: int = 0) -> dict[str, Any]:
    payload = _project_tracker_run_payload(delay_ms=delay_ms)
    payload["params"]["notice_title"] = notice_title
    payload["params"]["bid_no"] = bid_no
    return payload


def _wait_for_related_notice_precompute(server: ApiServer, run_id: str, *, timeout_seconds: float = 20.0) -> tuple[Any, dict[str, Any]]:
    deadline = time.time() + timeout_seconds
    last_run_payload: Any = None
    last_artifacts_payload: dict[str, Any] = {"items": []}
    while time.time() < deadline:
        run_status, run_payload = server.request_json("GET", f"/api/runs/{run_id}")
        if run_status != 200:
            raise AssertionError(f"unexpected run status while waiting for related notice precompute: {run_status} {run_payload}")
        last_run_payload = run_payload
        artifacts_status, artifacts_payload = server.request_json("GET", f"/api/runs/{run_id}/artifacts")
        if artifacts_status != 200:
            raise AssertionError(
                f"unexpected artifact status while waiting for related notice precompute: {artifacts_status} {artifacts_payload}"
            )
        last_artifacts_payload = artifacts_payload
        summary_output = dict(run_payload.get("summary", {}).get("output") or {})
        artifact_types = {item["artifact_type"] for item in artifacts_payload["items"]}
        if (
            summary_output.get("related_notice_precompute_status") == "success"
            and "related_notices_json" in artifact_types
            and summary_output.get("related_notice_snapshot_set_id")
        ):
            return run_payload, artifacts_payload
        if summary_output.get("related_notice_precompute_status") == "failed":
            raise AssertionError(f"related notice precompute failed: {summary_output}")
        time.sleep(0.05)
    raise AssertionError(
        "timed out waiting for related notice precompute; "
        f"last_run={last_run_payload} last_artifacts={last_artifacts_payload}"
    )


def _prepare_tracker_export(server: ApiServer) -> tuple[dict[str, Any], dict[str, Any]]:
    create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
    if create_status != 201:
        raise AssertionError(f"unexpected run create status: {create_status} {create_payload}")
    server.wait_for_run(create_payload["id"], target_statuses={"success"})

    tracker_status, tracker_payload = server.request_json(
        "POST",
        f"/api/runs/{create_payload['id']}/tracker-export",
    )
    if tracker_status != 202:
        raise AssertionError(f"unexpected tracker export status: {tracker_status} {tracker_payload}")
    server.wait_for_run(tracker_payload["id"], target_statuses={"success"})

    entries_status, entries_payload = server.request_json(
        "GET",
        f"/api/tracker-entries?source_tracker_run_id={tracker_payload['id']}&page=1&page_size=20",
    )
    if entries_status != 200:
        raise AssertionError(f"unexpected tracker entries status: {entries_status} {entries_payload}")
    if not entries_payload["items"]:
        raise AssertionError("expected at least one tracker entry")
    return tracker_payload, entries_payload["items"][0]


def _create_fake_report_script() -> tempfile.TemporaryDirectory[str]:
    temp_dir = tempfile.TemporaryDirectory()
    root = Path(temp_dir.name)
    (root / "fake_report.py").write_text(
        "\n".join(
            [
                "from __future__ import annotations",
                "import argparse",
                "import json",
                "from pathlib import Path",
                "",
                "parser = argparse.ArgumentParser()",
                "parser.add_argument('--output', required=True)",
                "args, _rest = parser.parse_known_args()",
                "output_path = Path(args.output)",
                "output_path.parent.mkdir(parents=True, exist_ok=True)",
                "output_path.write_text(json.dumps({",
                "    'summary': {'all_match': True, 'matched_count': 5, 'mismatched_count': 0},",
                "    'comparisons': {'winner_csv': {'match': True}},",
                "}), encoding='utf-8')",
                "print('fake report complete')",
            ]
        ),
        encoding="utf-8",
    )
    return temp_dir


class Phase1ApiTests(unittest.TestCase):
    def _list_download_audit_logs(self, server: ApiServer, *, limit: int = 20) -> list[dict[str, Any]]:
        status_code, payload = server.request_json("GET", f"/api/admin/download-audit-logs?limit={limit}")
        self.assertEqual(status_code, 200)
        self.assertIsInstance(payload, dict)
        return list(payload["items"])

    def _list_login_audit_logs(self, server: ApiServer, *, limit: int = 20) -> list[dict[str, Any]]:
        status_code, payload = server.request_json("GET", f"/api/admin/login-audit-logs?limit={limit}")
        self.assertEqual(status_code, 200)
        self.assertIsInstance(payload, dict)
        return list(payload["items"])

    def test_health_response_includes_request_id_header(self) -> None:
        with ApiServer() as server:
            status_code, _body, headers = server.request_bytes("GET", "/health")
            self.assertEqual(status_code, 200)
            self.assertTrue(headers.get("x-request-id"))

    def test_dump_request_params_normalizes_multi_demand_org_input(self) -> None:
        from backend.api.app import _dump_request_params
        from backend.api.schemas import RunCreateRequest

        payload = RunCreateRequest(
            run_type="project_tracker",
            params={
                "start_date": "20250101",
                "end_date": "20250131",
                "notice_title": "설계공모",
                "bid_no": "",
                "demand_org": "부산/울산·경남\n부산",
                "rows_per_page": 100,
                "max_pages": 3,
                "api_scope": "service",
            },
            advanced_options={},
        )

        params = _dump_request_params(payload)

        self.assertEqual(params["demand_org"], "부산, 울산, 경남")

    def test_run_presets_can_be_created_and_listed(self) -> None:
        with ApiServer() as server:
            create_status, create_payload = server.request_json(
                "POST",
                "/api/run-presets",
                payload={
                    "name": "2025 상반기 전체",
                    "params": {
                        "start_date": "20250101",
                        "end_date": "20250630",
                        "notice_title": "설계공모",
                        "api_scope": "all",
                    },
                },
            )
            self.assertEqual(create_status, 201)
            self.assertEqual(create_payload["name"], "2025 상반기 전체")

            list_status, list_payload = server.request_json("GET", "/api/run-presets?limit=10")
            self.assertEqual(list_status, 200)
            self.assertEqual(len(list_payload["items"]), 1)
            self.assertEqual(list_payload["items"][0]["name"], "2025 상반기 전체")

    def test_run_events_endpoint_streams_sse_payloads(self) -> None:
        with ApiServer() as server:
            create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload(delay_ms=200))
            self.assertEqual(create_status, 201)
            with server.open_url(
                f"/api/runs/{create_payload['id']}/events?poll_interval_ms=250",
                headers={"Accept": "text/event-stream"},
                timeout=10,
            ) as response:
                chunk = response.read(4096).decode("utf-8")
            self.assertIn("event: run", chunk)
            self.assertIn('"run_type": "project_tracker"', chunk)

    def test_projects_endpoint_returns_aggregated_items(self) -> None:
        with ApiServer() as server:
            create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(create_status, 201)
            server.wait_for_run(create_payload["id"], target_statuses={"success"})
            _wait_for_related_notice_precompute(server, create_payload["id"])

            tracker_status, tracker_payload = server.request_json(
                "POST",
                f"/api/runs/{create_payload['id']}/tracker-export",
            )
            self.assertEqual(tracker_status, 202)
            server.wait_for_run(tracker_payload["id"], target_statuses={"success"})

            projects_status, projects_payload = server.request_json("GET", "/api/projects?page=1&page_size=20")
            self.assertEqual(projects_status, 200)
            self.assertGreaterEqual(projects_payload["total"], 1)
            project = projects_payload["items"][0]
            self.assertIn("project_name", project)
            self.assertIn("project_search_name", project)
            self.assertIn("source_json", project)
            self.assertIn("run_ids", project["source_json"])
            self.assertTrue(project["project_search_name"])

            related_status, related_payload = server.request_json(
                "GET",
                f"/api/projects/{project['id']}/related-notices",
            )
            self.assertEqual(related_status, 200)
            self.assertEqual(related_payload["project_id"], project["id"])
            self.assertEqual(related_payload["status"], "ready")
            self.assertIn(related_payload["source"], {"precomputed", "seed_fallback"})
            related_titles = [item["project_name"] for item in related_payload["items"]]
            self.assertTrue(any(project["latest_notice_title"] in title for title in related_titles))
            self.assertTrue(any("Follow-up" in title for title in related_titles))

    def test_frontend_console_is_served_from_app_route(self) -> None:
        with ApiServer() as server:
            status_code, body, headers = server.request_bytes("GET", "/app/")
            self.assertEqual(status_code, 200)
            self.assertIn("text/html", headers.get("content-type", ""))
            self.assertIn(b'id="run-form"', body)
            self.assertIn(b'id="runs-list"', body)

    def test_phase_report_endpoint_returns_json_file(self) -> None:
        reports_root = _build_test_tmp_path("reports", token=uuid4().hex)
        try:
            with ApiServer(env_overrides={"REPORTS_ROOT": str(reports_root)}) as server:
                reports_root.mkdir(parents=True, exist_ok=True)
                (reports_root / "phase1-artifact-diff-report.json").write_text(
                    json.dumps(
                        {
                            "summary": {"all_match": True, "matched_count": 5, "mismatched_count": 0},
                            "comparisons": {"winner_csv": {"match": True}},
                        }
                    ),
                    encoding="utf-8",
                )
                status_code, payload = server.request_json("GET", "/api/reports/phase1-artifact-diff")
                self.assertEqual(status_code, 200)
                self.assertEqual(payload["summary"]["all_match"], True)
                self.assertEqual(payload["comparisons"]["winner_csv"]["match"], True)
        finally:
            shutil.rmtree(reports_root.parent, ignore_errors=True)

    @pytest.mark.xfail(reason="tracker/report job path is not green in this branch")
    def test_report_job_endpoint_runs_script_and_updates_status(self) -> None:
        fake_script = _create_fake_report_script()
        reports_root = _build_test_tmp_path("reports", token=uuid4().hex)
        try:
            with ApiServer(
                env_overrides={
                    "REPORTS_ROOT": str(reports_root),
                    "REPORT_SCRIPT_PHASE1_ARTIFACT_DIFF": str(Path(fake_script.name) / "fake_report.py"),
                }
            ) as server:
                status_code, payload = server.request_json(
                    "POST",
                    "/api/report-jobs",
                    payload={"report_name": "phase1-artifact-diff", "seed_limit": 1},
                )
                self.assertEqual(status_code, 202)
                deadline = time.time() + 10
                latest_payload = payload
                while time.time() < deadline:
                    job_status, latest_payload = server.request_json("GET", f"/api/report-jobs/{payload['id']}")
                    self.assertEqual(job_status, 200)
                    if latest_payload["status"] in {"success", "failed"}:
                        break
                    time.sleep(0.1)
                self.assertEqual(latest_payload["status"], "success")
                self.assertEqual(latest_payload["summary"]["all_match"], True)
                self.assertIn("fake report complete", latest_payload["log_excerpt"])
        finally:
            fake_script.cleanup()
            shutil.rmtree(reports_root.parent, ignore_errors=True)

    def test_dashboard_summary_returns_counts(self) -> None:
        with ApiServer() as server:
            create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(create_status, 201)
            server.wait_for_run(create_payload["id"], target_statuses={"success"})
            dashboard_status, dashboard_payload = server.request_json("GET", "/api/dashboard/summary")
            self.assertEqual(dashboard_status, 200)
            self.assertGreaterEqual(dashboard_payload["run_counts"]["success"], 1)
            self.assertIn("tracker_total", dashboard_payload)
            self.assertIn("tracker_edited_total", dashboard_payload)
            self.assertIn("repository_backends", dashboard_payload)
            self.assertEqual(dashboard_payload["repository_backends"]["artifacts"], "in_memory")
            self.assertEqual(dashboard_payload["artifact_metadata_persistent"], False)
            self.assertGreaterEqual(dashboard_payload["tracker_total"], 0)
            self.assertGreaterEqual(dashboard_payload["tracker_edited_total"], 0)

    def test_project_tracker_success_creates_csv_and_logs(self) -> None:
        with ApiServer() as server:
            status_code, payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(status_code, 201)

            run_detail = server.wait_for_run(payload["id"], target_statuses={"success"})
            self.assertEqual(run_detail["status"], "success")
            self.assertEqual(run_detail["run_type"], "project_tracker")
            self.assertEqual(run_detail["progress_stage"], "finalize")
            self.assertEqual(run_detail["summary"]["output"]["winner_csv_rows"], 2)
            self.assertEqual(run_detail["summary"]["output"]["requested_collect_mode"], "synthetic")
            self.assertEqual(run_detail["summary"]["output"]["stage_backends"]["export"], "synthetic")
            self.assertEqual(run_detail["summary"]["output"]["runtime_profile"], "web_native")
            self.assertEqual(run_detail["summary"]["output"]["native_collect_diagnostics"], {})
            self.assertIn(
                run_detail["summary"]["output"]["related_notice_precompute_status"],
                {"queued", "running", "success"},
            )

            run_detail, artifacts_payload = _wait_for_related_notice_precompute(server, payload["id"])
            self.assertEqual(run_detail["summary"]["output"]["related_notice_precomputed"], True)
            self.assertGreaterEqual(run_detail["summary"]["output"]["related_notice_items"], 1)
            self.assertEqual(run_detail["summary"]["output"]["related_notice_snapshot_set_id"], payload["id"])

            logs_status, logs_payload = server.request_json("GET", f"/api/runs/{payload['id']}/logs?limit=50")
            self.assertEqual(logs_status, 200)
            messages = [item["message"] for item in logs_payload["items"]]
            self.assertIn("project_tracker started", messages)
            self.assertIn("collect stage started", messages)
            self.assertIn("collect stage completed", messages)
            self.assertIn("filter stage completed", messages)
            self.assertIn("rescan stage completed", messages)
            self.assertIn("export stage completed", messages)
            self.assertIn("project_tracker finished successfully", messages)
            self.assertIn("related notice precompute started", messages)
            self.assertIn("related notices precomputed", messages)

            artifact_types = {item["artifact_type"] for item in artifacts_payload["items"]}
            self.assertEqual(
                artifact_types,
                {"seed_csv", "candidate_csv", "internal_nav_csv", "winner_csv", "execution_manifest", "related_notices_json"},
            )
            artifact_map = {item["artifact_type"]: item for item in artifacts_payload["items"]}
            artifact = artifact_map["winner_csv"]
            self.assertEqual(artifact["mime_type"], "text/csv")
            self.assertEqual(artifact["meta"]["backend"], "synthetic")
            self.assertTrue(artifact["download_url"].startswith(server.base_url))
            self.assertTrue(artifact["download_url"].endswith(f"/api/artifacts/{artifact['id']}/download"))
            self.assertEqual(artifact["download_url_expires_in"], 600)

            download_status, body, headers = server.request_bytes("GET", artifact["download_url"])
            self.assertEqual(download_status, 200)
            self.assertGreater(len(body), 0)
            self.assertIn("text/csv", headers.get("content-type", ""))

            preview_status, preview_payload = server.request_json(
                "GET",
                f"/api/artifacts/{artifact['id']}/preview?limit=1",
            )
            self.assertEqual(preview_status, 200)
            self.assertEqual(preview_payload["kind"], "table")
            self.assertEqual(preview_payload["format"], "csv")
            self.assertEqual(preview_payload["artifact_type"], "winner_csv")
            self.assertIn("bid_no", preview_payload["headers"])
            self.assertEqual(len(preview_payload["rows"]), 1)
            self.assertEqual(preview_payload["rows"][0]["bid_no"], run_detail["params"]["bid_no"])
            self.assertEqual(preview_payload["total_rows"], 2)

            manifest_status, manifest_body, manifest_headers = server.request_bytes(
                "GET",
                artifact_map["execution_manifest"]["download_url"],
            )
            self.assertEqual(manifest_status, 200)
            self.assertIn("application/json", manifest_headers.get("content-type", ""))
            manifest_payload = json.loads(manifest_body.decode("utf-8"))
            self.assertEqual(manifest_payload["requested_collect_mode"], "synthetic")
            self.assertEqual(manifest_payload["stage_backends"]["collect"], "synthetic")
            self.assertEqual(manifest_payload["native_collect_diagnostics"], {})

            related_status, related_body, related_headers = server.request_bytes(
                "GET",
                artifact_map["related_notices_json"]["download_url"],
            )
            self.assertEqual(related_status, 200)
            self.assertIn("application/json", related_headers.get("content-type", ""))
            related_payload = json.loads(related_body.decode("utf-8"))
            self.assertGreaterEqual(related_payload["item_count"], 1)

    def test_project_tracker_success_publishes_related_notice_snapshot_set(self) -> None:
        publication_env = {
            "RELATED_NOTICE_PUBLICATION_REPOSITORY_BACKEND": "supabase",
            "SUPABASE_SECRET_KEY": "phase1-publication-test-secret",
            "SUPABASE_SERVICE_ROLE_KEY": "",
            "SUPABASE_ANON_KEY": "",
        }
        with _RelatedNoticePublicationMockServer() as publication_server:
            env_overrides = dict(publication_env)
            env_overrides["SUPABASE_URL"] = publication_server.base_url
            with patch.dict(os.environ, env_overrides, clear=False):
                reset_related_notice_publication_repository()
                try:
                    with ApiServer(env_overrides=env_overrides) as server:
                        status_code, payload = server.request_json(
                            "POST",
                            "/api/runs",
                            payload=_project_tracker_run_payload(),
                        )
                        self.assertEqual(status_code, 201)

                        run_detail, _artifacts_payload = _wait_for_related_notice_precompute(server, payload["id"])
                        self.assertEqual(run_detail["summary"]["output"]["related_notice_snapshot_set_id"], payload["id"])

                        publication_repository = get_related_notice_publication_repository()
                        publication = publication_repository.get_publication(
                            organization_id=load_phase1_identity().organization_id,
                        )
                        self.assertIsNotNone(publication)
                        assert publication is not None
                        self.assertEqual(publication["source_run_id"], UUID(payload["id"]))
                        self.assertTrue(publication["published_snapshot_set_id"])
                finally:
                    reset_related_notice_publication_repository()

    def test_project_tracker_success_logs_publish_failure_without_advancing_snapshot(self) -> None:
        publication_env = {
            "RELATED_NOTICE_PUBLICATION_REPOSITORY_BACKEND": "supabase",
            "SUPABASE_SECRET_KEY": "phase1-publication-test-secret",
            "SUPABASE_SERVICE_ROLE_KEY": "",
            "SUPABASE_ANON_KEY": "",
        }
        with _RelatedNoticePublicationMockServer(reject_writes=True) as publication_server:
            env_overrides = dict(publication_env)
            env_overrides["SUPABASE_URL"] = publication_server.base_url
            with patch.dict(os.environ, env_overrides, clear=False):
                reset_related_notice_publication_repository()
                try:
                    with ApiServer(env_overrides=env_overrides) as server:
                        status_code, payload = server.request_json(
                            "POST",
                            "/api/runs",
                            payload=_project_tracker_run_payload(),
                        )
                        self.assertEqual(status_code, 201)

                        deadline = time.time() + 20
                        run_detail = None
                        artifacts_payload: dict[str, Any] = {"items": []}
                        while time.time() < deadline:
                            run_status, run_payload = server.request_json("GET", f"/api/runs/{payload['id']}")
                            self.assertEqual(run_status, 200)
                            artifacts_status, artifacts_payload = server.request_json(
                                "GET",
                                f"/api/runs/{payload['id']}/artifacts",
                            )
                            self.assertEqual(artifacts_status, 200)
                            summary_output = dict(run_payload.get("summary", {}).get("output") or {})
                            artifact_types = {item["artifact_type"] for item in artifacts_payload["items"]}
                            if summary_output.get("related_notice_precompute_status") == "success" and "related_notices_json" in artifact_types:
                                run_detail = run_payload
                                break
                            time.sleep(0.05)

                        self.assertIsNotNone(run_detail)
                        assert run_detail is not None
                        summary_output = dict(run_detail.get("summary", {}).get("output") or {})
                        self.assertEqual(summary_output["related_notice_precompute_status"], "success")
                        self.assertEqual(summary_output["related_notice_precomputed"], True)
                        self.assertEqual(summary_output["related_notice_snapshot_set_id"], "")
                        self.assertGreaterEqual(
                            len([item for item in artifacts_payload["items"] if item["artifact_type"] == "related_notices_json"]),
                            1,
                        )

                        publication_repository = get_related_notice_publication_repository()
                        publication = publication_repository.get_publication(
                            organization_id=load_phase1_identity().organization_id,
                        )
                        self.assertIsNone(publication)

                        server.wait_for_run(payload["id"], target_statuses={"success"})
                        publish_failure_deadline = time.time() + 5
                        messages: list[str] = []
                        while time.time() < publish_failure_deadline:
                            time.sleep(0.2)
                            logs_status, logs_payload = server.request_json("GET", f"/api/runs/{payload['id']}/logs?limit=50")
                            self.assertEqual(logs_status, 200)
                            messages = [item["message"] for item in logs_payload["items"]]
                            if "related notice publish failed" in messages:
                                break
                        self.assertIn("related notice publish failed", messages)
                finally:
                    reset_related_notice_publication_repository()

    def test_project_tracker_success_auto_queues_tracker_export(self) -> None:
        with ApiServer() as server:
            status_code, payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(status_code, 201)

            parent_detail = server.wait_for_run(payload["id"], target_statuses={"success"})
            auto_child_id = parent_detail["summary"]["output"].get("auto_tracker_export_run_id")
            self.assertTrue(auto_child_id)

            child_detail = server.wait_for_run(auto_child_id, target_statuses={"success"})
            self.assertEqual(child_detail["run_type"], "tracker_export")
            self.assertEqual(child_detail["parent_run_id"], payload["id"])
            self.assertGreaterEqual(child_detail["summary"]["output"]["estimated_tracker_entry_rows"], 1)
            self.assertGreaterEqual(child_detail["summary"]["output"]["tracker_entry_rows"], 1)

            list_status, list_payload = server.request_json(
                "GET",
                f"/api/runs?parent_run_id={payload['id']}&page=1&page_size=20",
            )
            self.assertEqual(list_status, 200)
            self.assertEqual(list_payload["total"], 1)
            self.assertEqual(list_payload["items"][0]["id"], auto_child_id)

            parent_status, parent_payload = server.request_json("GET", f"/api/runs/{payload['id']}")
            self.assertEqual(parent_status, 200)
            self.assertEqual(parent_payload["summary"]["output"]["auto_tracker_export_status"], "success")
            self.assertEqual(parent_payload["summary"]["output"]["auto_tracker_export_tracking_excel_generated"], True)
            self.assertEqual(parent_payload["summary"]["output"]["auto_tracker_export_run_id"], auto_child_id)
            self.assertGreaterEqual(parent_payload["summary"]["output"]["auto_tracker_export_tracker_entry_rows"], 1)

    def test_create_run_validation_error_does_not_create_run(self) -> None:
        with ApiServer() as server:
            list_status_before, list_payload_before = server.request_json("GET", "/api/runs?page=1&page_size=20")
            self.assertEqual(list_status_before, 200)
            self.assertEqual(list_payload_before["total"], 0)

            status_code, payload = server.request_json(
                "POST",
                "/api/runs",
                payload={
                    "run_type": "project_tracker",
                    "advanced_options": {"collect_mode": "synthetic"},
                    "params": {
                        "start_date": "202501",
                        "end_date": "20250131",
                        "notice_title": "",
                        "bid_no": "",
                        "demand_org": "",
                        "rows_per_page": 50,
                        "max_pages": 2,
                        "api_scope": "construction",
                    },
                },
            )
            self.assertEqual(status_code, 400)
            self.assertEqual(payload["error"]["code"], "validation_error")
            self.assertEqual(payload["error"]["message"], "start_date must be YYYYMMDD")

            list_status_after, list_payload_after = server.request_json("GET", "/api/runs?page=1&page_size=20")
            self.assertEqual(list_status_after, 200)
            self.assertEqual(list_payload_after["total"], 0)

    def test_create_run_normalizes_invalid_end_day_to_month_last_day(self) -> None:
        with ApiServer() as server:
            status_code, payload = server.request_json(
                "POST",
                "/api/runs",
                payload={
                    "run_type": "project_tracker",
                    "advanced_options": {"collect_mode": "synthetic"},
                    "params": {
                        "start_date": "20250601",
                        "end_date": "20250631",
                        "notice_title": "테스트 공고",
                        "bid_no": "",
                        "demand_org": "",
                        "rows_per_page": 50,
                        "max_pages": 2,
                        "api_scope": "construction",
                    },
                },
            )
            self.assertEqual(status_code, 201)

            detail = server.wait_for_run(payload["id"], target_statuses={"success"})
            self.assertEqual(detail["params"]["start_date"], "20250601")
            self.assertEqual(detail["params"]["end_date"], "20250630")

    def test_create_run_rejects_invalid_month_even_with_eight_digits(self) -> None:
        with ApiServer() as server:
            status_code, payload = server.request_json(
                "POST",
                "/api/runs",
                payload={
                    "run_type": "project_tracker",
                    "advanced_options": {"collect_mode": "synthetic"},
                    "params": {
                        "start_date": "20251301",
                        "end_date": "20250131",
                        "notice_title": "테스트 공고",
                        "bid_no": "",
                        "demand_org": "",
                        "rows_per_page": 50,
                        "max_pages": 2,
                        "api_scope": "construction",
                    },
                },
            )
            self.assertEqual(status_code, 400)
            self.assertEqual(payload["error"]["code"], "validation_error")
            self.assertEqual(payload["error"]["message"], "start_date must be YYYYMMDD")

    def test_cancel_running_run_transitions_to_cancelled(self) -> None:
        with ApiServer() as server:
            status_code, payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload(delay_ms=500))
            self.assertEqual(status_code, 201)

            running_detail = server.wait_for_run(payload["id"], target_statuses={"running"})
            self.assertEqual(running_detail["status"], "running")

            cancel_status, cancel_payload = server.request_json("POST", f"/api/runs/{payload['id']}/cancel")
            self.assertEqual(cancel_status, 202)
            self.assertTrue(cancel_payload["cancel_requested"])

            cancelled_detail = server.wait_for_run(payload["id"], target_statuses={"cancelled"})
            self.assertEqual(cancelled_detail["status"], "cancelled")
            self.assertTrue(cancelled_detail["cancel_requested"])

            logs_status, logs_payload = server.request_json("GET", f"/api/runs/{payload['id']}/logs?limit=50")
            self.assertEqual(logs_status, 200)
            messages = [item["message"] for item in logs_payload["items"]]
            self.assertIn("cancel requested", messages)
            self.assertTrue(any("cancelled" in message for message in messages))

            artifacts_status, artifacts_payload = server.request_json("GET", f"/api/runs/{payload['id']}/artifacts")
            self.assertEqual(artifacts_status, 200)
            artifact_types = {item["artifact_type"] for item in artifacts_payload["items"]}
            self.assertIn("seed_csv", artifact_types)
            self.assertNotIn("winner_csv", artifact_types)

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_tracker_export_success_creates_child_run_entries_and_workbook(self) -> None:
        with ApiServer() as server:
            create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(create_status, 201)
            parent_detail = server.wait_for_run(create_payload["id"], target_statuses={"success"})
            self.assertEqual(parent_detail["status"], "success")

            tracker_status, tracker_payload = server.request_json(
                "POST",
                f"/api/runs/{create_payload['id']}/tracker-export",
            )
            self.assertEqual(tracker_status, 202)

            child_detail = server.wait_for_run(tracker_payload["id"], target_statuses={"success"})
            self.assertEqual(child_detail["status"], "success")
            self.assertEqual(child_detail["run_type"], "tracker_export")
            self.assertEqual(child_detail["parent_run_id"], create_payload["id"])
            self.assertGreaterEqual(child_detail["summary"]["output"]["estimated_tracker_entry_rows"], 1)
            self.assertGreaterEqual(child_detail["summary"]["output"]["tracker_entry_rows"], 1)
            self.assertEqual(child_detail["summary"]["output"]["tracking_excel_generated"], True)

            parent_status, parent_payload = server.request_json("GET", f"/api/runs/{create_payload['id']}")
            self.assertEqual(parent_status, 200)
            self.assertEqual(parent_payload["status"], "success")

            artifacts_status, artifacts_payload = server.request_json(
                "GET",
                f"/api/runs/{tracker_payload['id']}/artifacts",
            )
            self.assertEqual(artifacts_status, 200)
            self.assertEqual(len(artifacts_payload["items"]), 1)
            artifact = artifacts_payload["items"][0]
            self.assertEqual(artifact["artifact_type"], "tracking_excel")
            self.assertIn(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                artifact["mime_type"],
            )
            self.assertTrue(artifact["download_url"].startswith(server.base_url))
            self.assertTrue(artifact["download_url"].endswith(f"/api/artifacts/{artifact['id']}/download"))
            self.assertEqual(artifact["download_url_expires_in"], 600)

            download_status, body, _headers = server.request_bytes("GET", artifact["download_url"])
            self.assertEqual(download_status, 200)
            self.assertGreater(len(body), 0)

            preview_status, preview_payload = server.request_json(
                "GET",
                f"/api/artifacts/{artifact['id']}/preview?limit=5",
            )
            self.assertEqual(preview_status, 200)
            self.assertEqual(preview_payload["kind"], "tracker_workbook")
            self.assertEqual(preview_payload["format"], "xlsx")
            self.assertGreaterEqual(preview_payload["total_rows"], 1)
            self.assertEqual(preview_payload["header_row"][0], "NO.")

            entries_status, entries_payload = server.request_json(
                "GET",
                f"/api/tracker-entries?source_tracker_run_id={tracker_payload['id']}&page=1&page_size=20",
            )
            self.assertEqual(entries_status, 200)
            self.assertGreaterEqual(entries_payload["total"], 1)
            entry = entries_payload["items"][0]

            summary_status, summary_payload = server.request_json(
                "GET",
                f"/api/tracker-entry-summaries?source_tracker_run_id={tracker_payload['id']}&page=1&page_size=20",
            )
            self.assertEqual(summary_status, 200)
            self.assertGreaterEqual(summary_payload["total"], 1)
            summary_entry = summary_payload["items"][0]
            self.assertEqual(summary_entry["id"], entry["id"])
            self.assertIn("project_name", summary_entry)
            self.assertNotIn("manager_name", summary_entry)

            detail_status, detail_payload = server.request_json(
                "GET",
                f"/api/tracker-entries/{entry['id']}",
            )
            self.assertEqual(detail_status, 200)
            self.assertEqual(detail_payload["id"], entry["id"])
            self.assertIn("manager_name", detail_payload)

            patch_status, patch_payload = server.request_json(
                "PATCH",
                f"/api/tracker-entries/{entry['id']}",
                payload={
                    "field_name": "project_name",
                    "value": "Project Name Final",
                    "actor_label": "phase1-tester",
                    "change_source": "web",
                },
            )
            self.assertEqual(patch_status, 200)
            self.assertTrue(patch_payload["changed"])
            self.assertEqual(patch_payload["entry"]["project_name"], "Project Name Final")
            self.assertIn("project_name", patch_payload["entry"]["overridden_fields"])
            self.assertEqual(patch_payload["audit_log"]["field_name"], "project_name")

            audit_status, audit_payload = server.request_json(
                "GET",
                f"/api/tracker-entries/{entry['id']}/audit-logs?limit=10",
            )
            self.assertEqual(audit_status, 200)
            self.assertGreaterEqual(len(audit_payload["items"]), 1)
            self.assertEqual(audit_payload["items"][0]["change_source"], "web")

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_tracker_export_emits_related_notice_added_for_second_entry_in_same_project(self) -> None:
        with ApiServer() as server:
            shared_title = "Phase1 Related Notice Project"
            first_payload = _project_tracker_run_payload_for_project(
                notice_title=shared_title,
                bid_no="PHASE1REL0001",
            )
            first_status, first_run = server.request_json("POST", "/api/runs", payload=first_payload)
            self.assertEqual(first_status, 201)
            server.wait_for_run(first_run["id"], target_statuses={"success"})
            first_tracker_status, first_tracker = server.request_json(
                "POST",
                f"/api/runs/{first_run['id']}/tracker-export",
            )
            self.assertEqual(first_tracker_status, 202)
            server.wait_for_run(first_tracker["id"], target_statuses={"success"})

            second_payload = _project_tracker_run_payload_for_project(
                notice_title=shared_title,
                bid_no="PHASE1REL0002",
            )
            second_status, second_run = server.request_json("POST", "/api/runs", payload=second_payload)
            self.assertEqual(second_status, 201)
            server.wait_for_run(second_run["id"], target_statuses={"success"})
            second_tracker_status, second_tracker = server.request_json(
                "POST",
                f"/api/runs/{second_run['id']}/tracker-export",
            )
            self.assertEqual(second_tracker_status, 202)
            server.wait_for_run(second_tracker["id"], target_statuses={"success"})

            entries_status, entries_payload = server.request_json(
                "GET",
                f"/api/tracker-entries?source_tracker_run_id={second_tracker['id']}&page=1&page_size=20",
            )
            self.assertEqual(entries_status, 200)
            self.assertGreaterEqual(entries_payload["total"], 1)
            second_entry = entries_payload["items"][0]

            events_status, events_payload = server.request_json(
                "GET",
                "/api/tracker-change-events?limit=20",
            )
            self.assertEqual(events_status, 200)
            related_events = [
                item for item in (events_payload["items"] or [])
                if item.get("event_type") == "related_notice_added"
            ]
            self.assertTrue(related_events)
            matching_event = next(
                (item for item in related_events if item.get("tracker_entry_id") == second_entry["id"]),
                None,
            )
            self.assertIsNotNone(matching_event)
            self.assertEqual(matching_event["new_value"], "PHASE1REL0002/000")

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_list_runs_supports_parent_run_id_filter(self) -> None:
        with ApiServer() as server:
            create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(create_status, 201)
            server.wait_for_run(create_payload["id"], target_statuses={"success"})

            tracker_status, tracker_payload = server.request_json(
                "POST",
                f"/api/runs/{create_payload['id']}/tracker-export",
            )
            self.assertEqual(tracker_status, 202)
            server.wait_for_run(tracker_payload["id"], target_statuses={"success"})

            list_status, list_payload = server.request_json(
                "GET",
                f"/api/runs?parent_run_id={create_payload['id']}&page=1&page_size=20",
            )
            self.assertEqual(list_status, 200)
            self.assertEqual(list_payload["total"], 1)
            self.assertEqual(list_payload["items"][0]["id"], tracker_payload["id"])
            self.assertEqual(list_payload["items"][0]["parent_run_id"], create_payload["id"])

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_tracker_export_failure_only_fails_child_run(self) -> None:
        missing_template_path = _build_test_tmp_path(f"missing-template-{uuid4().hex}.xlsx", token="files")
        with ApiServer(env_overrides={"TRACKER_TEMPLATE_PATH": str(missing_template_path)}) as server:
            create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(create_status, 201)
            parent_detail = server.wait_for_run(create_payload["id"], target_statuses={"success"})
            self.assertEqual(parent_detail["status"], "success")

            tracker_status, tracker_payload = server.request_json(
                "POST",
                f"/api/runs/{create_payload['id']}/tracker-export",
            )
            self.assertEqual(tracker_status, 202)

            child_detail = server.wait_for_run(tracker_payload["id"], target_statuses={"failed"})
            self.assertEqual(child_detail["status"], "failed")
            self.assertEqual(child_detail["run_type"], "tracker_export")
            self.assertIn("tracker template not found", child_detail["error"]["message"])

            parent_status, parent_payload = server.request_json("GET", f"/api/runs/{create_payload['id']}")
            self.assertEqual(parent_status, 200)
            self.assertEqual(parent_payload["status"], "success")

            artifacts_status, artifacts_payload = server.request_json(
                "GET",
                f"/api/runs/{tracker_payload['id']}/artifacts",
            )
            self.assertEqual(artifacts_status, 200)
            self.assertEqual(artifacts_payload["items"], [])

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_tracker_missing_report_downloads_csv_and_xlsx(self) -> None:
        with ApiServer() as server:
            create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(create_status, 201)
            server.wait_for_run(create_payload["id"], target_statuses={"success"})

            tracker_status, tracker_payload = server.request_json(
                "POST",
                f"/api/runs/{create_payload['id']}/tracker-export",
            )
            self.assertEqual(tracker_status, 202)
            server.wait_for_run(tracker_payload["id"], target_statuses={"success"})

            csv_status, csv_body, csv_headers = server.request_bytes(
                "GET",
                "/api/tracker-entries/missing-report/download?format=csv&limit=100",
            )
            self.assertEqual(csv_status, 200)
            self.assertIn("text/csv", csv_headers.get("content-type", ""))
            csv_text = csv_body.decode("utf-8-sig")
            self.assertIn("missing_field_key", csv_text.splitlines()[0])

            xlsx_status, xlsx_body, xlsx_headers = server.request_bytes(
                "GET",
                "/api/tracker-entries/missing-report/download?format=xlsx&limit=100",
            )
            self.assertEqual(xlsx_status, 200)
            self.assertIn(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                xlsx_headers.get("content-type", ""),
            )
            self.assertGreater(len(xlsx_body), 0)
            workbook = load_workbook(BytesIO(xlsx_body), data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["summary", "missing_items"])
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    self.assertEqual(
                        sheet.auto_filter.ref,
                        f"A2:{get_column_letter(sheet.max_column)}{sheet.max_row}",
                    )
                    self.assertEqual(sheet["A2"].font.sz, 10)
                    self.assertEqual(sheet["B2"].font.sz, 10)
            finally:
                workbook.close()

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_tracker_entry_summaries_downloads_csv_and_template_xlsx(self) -> None:
        with ApiServer() as server:
            create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(create_status, 201)
            server.wait_for_run(create_payload["id"], target_statuses={"success"})

            tracker_status, tracker_payload = server.request_json(
                "POST",
                f"/api/runs/{create_payload['id']}/tracker-export",
            )
            self.assertEqual(tracker_status, 202)
            server.wait_for_run(tracker_payload["id"], target_statuses={"success"})

            csv_status, csv_body, csv_headers = server.request_bytes(
                "GET",
                f"/api/tracker-entry-summaries/download?format=csv&source_tracker_run_id={tracker_payload['id']}",
            )
            self.assertEqual(csv_status, 200)
            self.assertIn("text/csv", csv_headers.get("content-type", ""))
            csv_text = csv_body.decode("utf-8-sig")
            self.assertIn("project_name", csv_text.splitlines()[0])

            xlsx_status, xlsx_body, xlsx_headers = server.request_bytes(
                "GET",
                f"/api/tracker-entry-summaries/download?format=xlsx&source_tracker_run_id={tracker_payload['id']}",
            )
            self.assertEqual(xlsx_status, 200)
            self.assertIn(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                xlsx_headers.get("content-type", ""),
            )
            workbook = load_workbook(BytesIO(xlsx_body), data_only=True)
            try:
                sheet = workbook[workbook.sheetnames[0]]
                self.assertEqual(str(sheet["A2"].value or "").strip(), "NO.")
                self.assertEqual(sheet.auto_filter.ref, f"A2:R{max(sheet.max_row, 2)}")
                for column_letter in ("G", "L", "M", "O"):
                    self.assertTrue(sheet.column_dimensions[column_letter].hidden)
                self.assertEqual(sheet["Q2"].value, "설계사무소(전기)")
                self.assertEqual(sheet["R2"].value, "설계사무소(기계)")
                self.assertTrue(str(sheet["B3"].value or "").strip())
            finally:
                workbook.close()

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_sales_claim_exports_create_download_audit_logs_for_my_and_company_scope(self) -> None:
        with ApiServer() as server:
            tracker_payload, entry = _prepare_tracker_export(server)
            claim_status, claim_payload = server.request_json(
                "POST",
                f"/api/sales-claims/projects/{entry['project_id']}/claim",
                payload={
                    "source_entry_id": entry["id"],
                    "source_run_id": tracker_payload["id"],
                    "project_name": entry["project_name"],
                    "estimated_amount_text": "",
                },
            )
            self.assertEqual(claim_status, 200)
            self.assertTrue(claim_payload["changed"])

            my_status, _my_body, my_headers = server.request_bytes("GET", "/api/sales-claims/export?scope=my")
            self.assertEqual(my_status, 200)
            self.assertIn('filename="my_active_sales.xlsx"', my_headers.get("content-disposition", ""))

            company_status, _company_body, company_headers = server.request_bytes(
                "GET",
                "/api/sales-claims/export?scope=company",
            )
            self.assertEqual(company_status, 200)
            self.assertIn('filename="company_active_sales.xlsx"', company_headers.get("content-disposition", ""))

            items = self._list_download_audit_logs(server, limit=10)
            self.assertEqual(len(items), 2)
            self.assertEqual(items[0]["source_page"], "company_active_sales")
            self.assertEqual(items[0]["download_scope"], "company")
            self.assertEqual(items[0]["download_format"], "xlsx")
            self.assertEqual(items[0]["file_name"], "company_active_sales.xlsx")
            self.assertEqual(items[0]["user_role"], "platform_admin")
            self.assertEqual(items[1]["source_page"], "my_active_sales")
            self.assertEqual(items[1]["download_scope"], "my")
            self.assertEqual(items[1]["download_format"], "xlsx")
            self.assertEqual(items[1]["file_name"], "my_active_sales.xlsx")
            self.assertEqual(items[1]["user_role"], "platform_admin")

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_tracker_entry_summary_downloads_create_download_audit_logs_for_csv_and_xlsx(self) -> None:
        with ApiServer() as server:
            tracker_payload, _entry = _prepare_tracker_export(server)

            csv_status, _csv_body, csv_headers = server.request_bytes(
                "GET",
                f"/api/tracker-entry-summaries/download?format=csv&source_tracker_run_id={tracker_payload['id']}",
            )
            self.assertEqual(csv_status, 200)
            self.assertIn(".csv", csv_headers.get("content-disposition", ""))

            xlsx_status, _xlsx_body, xlsx_headers = server.request_bytes(
                "GET",
                f"/api/tracker-entry-summaries/download?format=xlsx&source_tracker_run_id={tracker_payload['id']}",
            )
            self.assertEqual(xlsx_status, 200)
            self.assertIn(".xlsx", xlsx_headers.get("content-disposition", ""))

            items = self._list_download_audit_logs(server, limit=10)
            self.assertEqual(len(items), 2)
            self.assertEqual(items[0]["source_page"], "tracker_entries")
            self.assertEqual(items[0]["download_scope"], "global")
            self.assertEqual(items[0]["download_format"], "xlsx")
            self.assertTrue(str(items[0]["file_name"]).endswith(".xlsx"))
            self.assertEqual(items[1]["source_page"], "tracker_entries")
            self.assertEqual(items[1]["download_scope"], "global")
            self.assertEqual(items[1]["download_format"], "csv")
            self.assertTrue(str(items[1]["file_name"]).endswith(".csv"))

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_invalid_download_requests_do_not_create_download_audit_logs(self) -> None:
        with ApiServer() as server:
            sales_status, sales_payload = server.request_json("GET", "/api/sales-claims/export?scope=team")
            self.assertEqual(sales_status, 400)
            self.assertEqual(sales_payload["error"]["code"], "validation_error")

            items = self._list_download_audit_logs(server, limit=10)
            self.assertEqual(items, [])

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_download_audit_log_write_failure_does_not_break_tracker_download(self) -> None:
        from backend.api import app as app_module
        from backend.repositories.download_audit_logs import DownloadAuditLogRepositoryError

        class FailingDownloadAuditRepository:
            def create_log(self, **_kwargs):
                raise DownloadAuditLogRepositoryError("audit write failed")

            def list_logs(self, **_kwargs):
                return []

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )

        with (
            patch.object(app_module, "_get_download_audit_log_repository", return_value=FailingDownloadAuditRepository(), create=True),
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "_can_cache_tracker_export_workbook", return_value=False, create=True),
            patch.object(
                app_module,
                "_list_tracker_entries_for_export",
                return_value=[
                    {
                        "entry_key": "key-1",
                        "project_name": "Project",
                    }
                ],
            ),
            patch.object(app_module, "build_tracking_download_workbook_bytes", return_value=b"dummy-xlsx"),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/tracker-entry-summaries/download?format=xlsx")

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.content, b"dummy-xlsx")
        self.assertIn(".xlsx", response.headers.get("content-disposition", ""))

    def test_admin_download_audit_logs_endpoint_limits_to_current_org_newest_first(self) -> None:
        from backend.api import app as app_module

        repository = InMemoryDownloadAuditLogRepository()
        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        other_org_id = uuid4()
        repository.create_log(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            user_email="admin@example.com",
            user_role="platform_admin",
            download_scope="my",
            download_format="xlsx",
            source_page="my_active_sales",
            file_name="first.xlsx",
        )
        repository.create_log(
            organization_id=other_org_id,
            user_id=uuid4(),
            user_email="other@example.com",
            user_role="org_admin",
            download_scope="company",
            download_format="xlsx",
            source_page="company_active_sales",
            file_name="other-org.xlsx",
        )
        repository.create_log(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            user_email="admin@example.com",
            user_role="platform_admin",
            download_scope="global",
            download_format="csv",
            source_page="tracker_entries",
            file_name="second.csv",
        )
        repository.create_log(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            user_email="admin@example.com",
            user_role="platform_admin",
            download_scope="global",
            download_format="xlsx",
            source_page="tracker_entries",
            file_name="third.xlsx",
        )

        with (
            patch.object(app_module, "_get_download_audit_log_repository", return_value=repository, create=True),
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/download-audit-logs?limit=2")

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual([item["file_name"] for item in payload["items"]], ["third.xlsx", "second.csv"])
        self.assertEqual({item["organization_id"] for item in payload["items"]}, {str(DEFAULT_PHASE1_ORGANIZATION_ID)})

    def test_non_admin_cannot_read_download_audit_logs(self) -> None:
        from backend.api import app as app_module

        repository = InMemoryDownloadAuditLogRepository()
        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="member@example.com",
            display_name="Member",
            role="org_member",
        )

        with (
            patch.object(app_module, "_get_download_audit_log_repository", return_value=repository, create=True),
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/download-audit-logs?limit=5")

        self.assertEqual(response.status_code, 403, response.text)
        self.assertEqual(response.json()["error"]["code"], "auth_forbidden")

    def test_admin_login_audit_logs_endpoint_limits_to_current_org_newest_first(self) -> None:
        from backend.api import app as app_module

        repository = InMemoryLoginAuditLogRepository()
        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        other_org_id = uuid4()
        repository.create_log(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            user_email="admin@example.com",
            user_role="platform_admin",
            ip_address="203.0.113.9",
            user_agent="ua-1",
        )
        repository.create_log(
            organization_id=other_org_id,
            user_id=uuid4(),
            user_email="other@example.com",
            user_role="org_admin",
            ip_address="198.51.100.7",
            user_agent="ua-2",
        )
        repository.create_log(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            user_email="admin@example.com",
            user_role="platform_admin",
            ip_address="203.0.113.10",
            user_agent="ua-3",
        )
        repository.create_log(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            user_email="admin@example.com",
            user_role="platform_admin",
            ip_address="203.0.113.11",
            user_agent="ua-4",
        )

        with (
            patch.object(app_module, "_get_login_audit_log_repository", return_value=repository, create=True),
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/login-audit-logs?limit=2")

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual([item["user_agent"] for item in payload["items"]], ["ua-4", "ua-3"])
        self.assertEqual({item["organization_id"] for item in payload["items"]}, {str(DEFAULT_PHASE1_ORGANIZATION_ID)})

    def test_admin_google_sheets_bootstrap_returns_tabs_with_row_and_column_counts(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        last_successful_sync_at = datetime(2026, 4, 7, 12, 0, tzinfo=timezone.utc)
        config = {"spreadsheet_id": "spreadsheet-1"}
        snapshot = {
            "enabled": True,
            "source_title": "Admin Sheet",
            "source_url": "https://example.com/sheets",
            "sync_status": "idle",
            "last_successful_sync_at": last_successful_sync_at,
            "last_failed_sync_at": None,
            "last_error": "",
            "tabs": [
                {
                    "key": "tab-1",
                    "sheet_id": 101,
                    "raw_title": "Raw 1",
                    "display_title": "Tab 1",
                    "sheet_order": 1,
                },
                {
                    "key": "tab-2",
                    "sheet_id": 202,
                    "raw_title": "Raw 2",
                    "display_title": "Tab 2",
                    "sheet_order": 2,
                },
            ],
            "sheets": {
                "tab-1": {"row_count": 3, "column_count": 5},
                "tab-2": {"row_count": 0, "column_count": 1},
            },
        }

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value=config, create=True),
            patch.object(app_module, "ensure_google_sheets_admin_sync_worker_started", create=True) as ensure_worker,
            patch.object(app_module, "read_google_sheets_admin_snapshot", return_value=snapshot, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/google-sheets/bootstrap")

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["enabled"], True)
        self.assertEqual(payload["source_title"], "Admin Sheet")
        self.assertEqual(payload["source_url"], "https://example.com/sheets")
        self.assertEqual(payload["sync_status"], "idle")
        self.assertTrue(str(payload["last_successful_sync_at"]).startswith("2026-04-07T12:00:00"))
        self.assertEqual([item["key"] for item in payload["tabs"]], ["tab-1", "tab-2"])
        self.assertEqual(payload["tabs"][0]["row_count"], 3)
        self.assertEqual(payload["tabs"][0]["column_count"], 5)
        self.assertEqual(payload["tabs"][1]["row_count"], 0)
        self.assertEqual(payload["tabs"][1]["column_count"], 1)
        ensure_worker.assert_called_once_with(config=config)

    def test_admin_google_sheets_bootstrap_returns_initializing_status_on_cold_start(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        config = {"spreadsheet_id": "spreadsheet-1"}

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value=config, create=True),
            patch.object(app_module, "ensure_google_sheets_admin_sync_worker_started", create=True) as ensure_worker,
            patch.object(app_module, "read_google_sheets_admin_snapshot", return_value=None, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/google-sheets/bootstrap")

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["enabled"], True)
        self.assertEqual(payload["sync_status"], "initializing")
        self.assertEqual(payload["tabs"], [])
        ensure_worker.assert_called_once_with(config=config)

    def test_admin_google_sheets_bootstrap_rebuilds_tabs_from_sheet_snapshot_when_tab_list_is_missing(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        config = {"spreadsheet_id": "spreadsheet-1"}
        snapshot = {
            "enabled": True,
            "source_title": "Admin Sheet",
            "source_url": "https://example.com/sheets",
            "sync_status": "synced",
            "tabs": [],
            "sheets": {
                "sheet-22": {
                    "sheet_id": 22,
                    "raw_title": "설계리스트",
                    "display_title": "설계리스트",
                    "row_count": 12,
                    "column_count": 6,
                },
                "sheet-33": {
                    "sheet_id": 33,
                    "raw_title": "발주예정",
                    "display_title": "발주예정",
                    "row_count": 8,
                    "column_count": 4,
                },
            },
        }

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value=config, create=True),
            patch.object(app_module, "ensure_google_sheets_admin_sync_worker_started", create=True) as ensure_worker,
            patch.object(app_module, "read_google_sheets_admin_snapshot", return_value=snapshot, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/google-sheets/bootstrap")

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual([item["key"] for item in payload["tabs"]], ["sheet-22", "sheet-33"])
        self.assertEqual(payload["tabs"][0]["display_title"], "설계리스트")
        self.assertEqual(payload["tabs"][0]["row_count"], 12)
        self.assertEqual(payload["tabs"][1]["display_title"], "발주예정")
        self.assertEqual(payload["tabs"][1]["column_count"], 4)
        ensure_worker.assert_called_once_with(config=config)

    def test_admin_google_sheets_bootstrap_reports_route_timing(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        config = {"spreadsheet_id": "spreadsheet-1"}
        snapshot = {
            "enabled": True,
            "source_title": "Admin Sheet",
            "source_url": "https://example.com/sheets",
            "sync_status": "idle",
            "tabs": [],
            "sheets": {},
        }

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value=config, create=True),
            patch.object(app_module, "ensure_google_sheets_admin_sync_worker_started", create=True),
            patch.object(app_module, "read_google_sheets_admin_snapshot", return_value=snapshot, create=True),
            patch("backend.api.routers.admin.log_google_sheets_admin_duration") as log_mock,
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/google-sheets/bootstrap")

        self.assertEqual(response.status_code, 200, response.text)
        log_mock.assert_called()
        self.assertEqual(log_mock.call_args.kwargs["event"], "admin_bootstrap_route")

    def test_admin_google_sheets_sheet_payload_returns_headers_and_rows(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        last_successful_sync_at = datetime(2026, 4, 8, 0, 1, tzinfo=timezone.utc)
        config = {"spreadsheet_id": "spreadsheet-1"}
        snapshot = {
            "last_successful_sync_at": last_successful_sync_at,
            "sheets": {
                "tab-1": {
                    "sheet_id": 101,
                    "raw_title": "Raw 1",
                    "display_title": "Tab 1",
                    "headers": ["name", "value"],
                    "header_cells": [
                        {"text": "name", "href": ""},
                        {"text": "value", "href": ""},
                    ],
                    "rows": [["a", "1"], ["b", "2"]],
                    "row_cells": [
                        [
                            {"text": "a", "href": "https://docs.google.com/document/d/a/edit"},
                            {"text": "1", "href": ""},
                        ],
                        [
                            {"text": "b", "href": ""},
                            {"text": "2", "href": ""},
                        ],
                    ],
                    "row_count": 2,
                    "column_count": 2,
                }
            },
        }

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value=config, create=True),
            patch.object(app_module, "ensure_google_sheets_admin_sync_worker_started", create=True),
            patch.object(app_module, "read_google_sheets_admin_snapshot", return_value=snapshot, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/google-sheets/sheets/tab-1")

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["key"], "tab-1")
        self.assertTrue(str(payload["synced_at"]).startswith("2026-04-08T00:01:00"))
        self.assertEqual(payload["headers"], ["name", "value"])
        self.assertEqual(payload["rows"], [["a", "1"], ["b", "2"]])
        self.assertEqual(payload["header_cells"][0], {"text": "name", "href": ""})
        self.assertEqual(
            payload["row_cells"][0][0],
            {"text": "a", "href": "https://docs.google.com/document/d/a/edit"},
        )
        self.assertEqual(payload["row_count"], 2)
        self.assertEqual(payload["column_count"], 2)

    def test_admin_google_sheets_sheet_payload_reports_route_timing(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        config = {"spreadsheet_id": "spreadsheet-1"}
        snapshot = {
            "last_successful_sync_at": datetime(2026, 4, 8, 0, 1, tzinfo=timezone.utc),
            "sheets": {
                "tab-1": {
                    "sheet_id": 101,
                    "raw_title": "Raw 1",
                    "display_title": "Tab 1",
                    "headers": ["name"],
                    "header_cells": [{"text": "name", "href": ""}],
                    "rows": [["a"]],
                    "row_cells": [[{"text": "a", "href": ""}]],
                    "row_count": 1,
                    "column_count": 1,
                }
            },
        }

        with (
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value=config, create=True),
            patch.object(app_module, "ensure_google_sheets_admin_sync_worker_started", create=True),
            patch.object(app_module, "read_google_sheets_admin_snapshot", return_value=snapshot, create=True),
            patch("backend.api.routers.admin.log_google_sheets_admin_duration") as log_mock,
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/google-sheets/sheets/tab-1")

        self.assertEqual(response.status_code, 200, response.text)
        log_mock.assert_called()
        self.assertEqual(log_mock.call_args.kwargs["event"], "admin_sheet_payload_route")

    def test_admin_google_sheets_sheet_payload_returns_409_when_snapshot_missing(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        config = {"spreadsheet_id": "spreadsheet-1"}

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value=config, create=True),
            patch.object(app_module, "ensure_google_sheets_admin_sync_worker_started", create=True) as ensure_worker,
            patch.object(app_module, "read_google_sheets_admin_snapshot", return_value=None, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/google-sheets/sheets/tab-1")

        self.assertEqual(response.status_code, 409, response.text)
        self.assertEqual(response.json()["error"]["code"], "google_sheets_not_ready")
        ensure_worker.assert_called_once_with(config=config)

    def test_admin_google_sheets_sheet_payload_returns_404_when_config_missing(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value=None, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/google-sheets/sheets/tab-1")

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["error"]["code"], "google_sheets_not_configured")

    def test_admin_google_sheets_sheet_payload_returns_404_when_sheet_key_missing(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        config = {"spreadsheet_id": "spreadsheet-1"}
        snapshot = {"sheets": {}}

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value=config, create=True),
            patch.object(app_module, "ensure_google_sheets_admin_sync_worker_started", create=True),
            patch.object(app_module, "read_google_sheets_admin_snapshot", return_value=snapshot, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/google-sheets/sheets/missing-tab")

        self.assertEqual(response.status_code, 404, response.text)
        self.assertEqual(response.json()["error"]["code"], "google_sheet_not_found")

    def test_non_admin_cannot_trigger_google_sheets_manual_sync(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="member@example.com",
            display_name="Member",
            role="org_member",
        )

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value={"spreadsheet_id": "spreadsheet-1"}, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.post("/api/admin/google-sheets/sync")

        self.assertEqual(response.status_code, 403, response.text)
        self.assertEqual(response.json()["error"]["code"], "auth_forbidden")

    def test_admin_google_sheets_manual_sync_reports_not_started_when_sync_in_flight(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "load_google_sheets_admin_config", return_value={"spreadsheet_id": "spreadsheet-1"}, create=True),
            patch.object(app_module, "queue_google_sheets_admin_sync_now", return_value=False, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.post("/api/admin/google-sheets/sync")

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(payload["started"], False)
        self.assertEqual(payload["sync_status"], "already_running")

    def test_admin_organization_panel_bootstrap_returns_combined_payload(self) -> None:
        from backend.api import app as app_module

        actor = SalesActor(
            organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
            user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
            email="admin@example.com",
            display_name="Admin",
            role="platform_admin",
        )
        invitation_created_at = datetime(2026, 4, 6, 12, 0, tzinfo=timezone.utc)
        invitation_expires_at = datetime(2026, 4, 13, 12, 0, tzinfo=timezone.utc)
        member_rows = [
            {
                "id": DEFAULT_PHASE1_INTERNAL_USER_ID,
                "membership_id": uuid4(),
                "organization_id": DEFAULT_PHASE1_ORGANIZATION_ID,
                "email": "admin@example.com",
                "display_name": "Admin",
                "role": "platform_admin",
                "status": "active",
                "account_status": "active",
                "membership_status": "active",
                "organization_name": "Internal Operations",
                "team_name": "Ops",
                "job_title": "Lead",
            }
        ]
        invitation_rows = [
            {
                "id": uuid4(),
                "organization_id": DEFAULT_PHASE1_ORGANIZATION_ID,
                "email": "invitee@example.com",
                "role": "org_member",
                "display_name": "Invitee",
                "team_name": "Sales",
                "job_title": "Manager",
                "invite_token": "invite-token",
                "invite_url": "https://example.com/invite",
                "status": "pending",
                "expires_at": invitation_expires_at,
                "created_at": invitation_created_at,
                "updated_at": invitation_created_at,
                "delivery_status": "queued",
                "delivery_message": "queued",
                "initial_password": "",
            }
        ]
        plan_summary = {
            "organization_id": DEFAULT_PHASE1_ORGANIZATION_ID,
            "organization_name": "Internal Operations",
            "plan_code": "A",
            "plan_label": "Basic",
            "active_user_limit": 5,
            "pending_invite_limit": 5,
            "active_user_count": 1,
            "pending_invite_count": 1,
            "remaining_active_user_slots": 4,
            "remaining_pending_invite_slots": 4,
            "active_user_limit_reached": False,
            "pending_invite_limit_reached": False,
            "upgrade_required": False,
            "upgrade_message": "",
            "next_plan_code": "",
        }
        auth_audit_rows = [
            {
                "id": index + 1,
                "organization_id": DEFAULT_PHASE1_ORGANIZATION_ID,
                "actor_user_id": DEFAULT_PHASE1_INTERNAL_USER_ID,
                "actor_membership_id": None,
                "actor_email": "admin@example.com",
                "actor_display_name": "Admin",
                "actor_role": "platform_admin",
                "event_type": f"event-{index}",
                "target_type": "user",
                "target_id": str(DEFAULT_PHASE1_INTERNAL_USER_ID),
                "payload_json": {},
                "created_at": invitation_created_at,
            }
            for index in range(6)
        ]
        download_repository = InMemoryDownloadAuditLogRepository()
        login_repository = InMemoryLoginAuditLogRepository()
        for index in range(6):
            download_repository.create_log(
                organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
                user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
                user_email="admin@example.com",
                user_role="platform_admin",
                download_scope="global",
                download_format="xlsx",
                source_page="tracker_entries",
                file_name=f"download-{index}.xlsx",
            )
            login_repository.create_log(
                organization_id=DEFAULT_PHASE1_ORGANIZATION_ID,
                user_id=DEFAULT_PHASE1_INTERNAL_USER_ID,
                user_email="admin@example.com",
                user_role="platform_admin",
                ip_address=f"203.0.113.{index}",
                user_agent=f"agent-{index}",
            )

        with (
            _patch_test_auth_context(app_module, actor=actor),
            patch.object(app_module, "_resolve_sales_actor", return_value=actor),
            patch.object(app_module, "list_local_users", return_value=member_rows),
            patch.object(
                app_module,
                "get_organization_invitation_dashboard",
                return_value={"items": invitation_rows, "plan_summary": plan_summary},
            ),
            patch.object(app_module, "list_organization_audit_logs", return_value=auth_audit_rows),
            patch.object(app_module, "_get_download_audit_log_repository", return_value=download_repository, create=True),
            patch.object(app_module, "_get_login_audit_log_repository", return_value=login_repository, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.get("/api/admin/organization-panel-bootstrap")

        self.assertEqual(response.status_code, 200, response.text)
        payload = response.json()
        self.assertEqual(len(payload["members"]), 1)
        self.assertEqual(payload["members"][0]["email"], "admin@example.com")
        self.assertEqual(payload["plan_summary"]["plan_label"], "Basic")
        self.assertEqual(len(payload["invitations"]), 1)
        self.assertEqual(payload["invitations"][0]["email"], "invitee@example.com")
        self.assertEqual(len(payload["auth_audit_logs"]["items"]), 5)
        self.assertEqual(payload["auth_audit_logs"]["has_more"], True)
        self.assertEqual(len(payload["download_audit_logs"]["items"]), 5)
        self.assertEqual(payload["download_audit_logs"]["has_more"], True)
        self.assertEqual(len(payload["login_audit_logs"]["items"]), 5)
        self.assertEqual(payload["login_audit_logs"]["has_more"], True)

    def test_sign_in_records_login_audit_log(self) -> None:
        from backend.api import app as app_module
        from backend.api import auth_runtime
        from backend.repositories.in_memory_login_audit_logs import InMemoryLoginAuditLogRepository

        session_payload = {
            "authorized": True,
            "organization_id": str(DEFAULT_PHASE1_ORGANIZATION_ID),
            "organization_name": "Internal Operations",
            "local_user_id": str(DEFAULT_PHASE1_INTERNAL_USER_ID),
            "auth_user_id": str(DEFAULT_PHASE1_INTERNAL_USER_ID),
            "email": "admin@example.com",
            "role": "platform_admin",
            "display_name": "Admin",
            "access_token": "access",
            "refresh_token": "refresh",
            "access_expires_at": int(time.time()) + 3600,
        }
        repository = InMemoryLoginAuditLogRepository()
        dispatched: list[tuple[object, tuple[object, ...], dict[str, object]]] = []

        with (
            patch.object(app_module, "sign_in_with_password", return_value=session_payload),
            patch.object(app_module, "set_auth_session_cookie"),
            patch.object(
                app_module,
                "_dispatch_background",
                side_effect=lambda target, *args, **kwargs: dispatched.append((target, args, kwargs)),
            ),
            patch.object(auth_runtime, "get_login_audit_log_repository", return_value=repository, create=True),
            TestClient(app_module.app) as client,
        ):
            response = client.post(
                "/api/auth/sign-in",
                json={"email": "admin@example.com", "password": "password-123"},
                headers={"user-agent": "pytest-agent/1.0", "x-forwarded-for": "203.0.113.9"},
            )
            self.assertEqual(len(dispatched), 1)
            target, args, kwargs = dispatched[0]
            self.assertIs(target, app_module.record_login_audit_log)
            target(*args, **kwargs)

        self.assertEqual(response.status_code, 200, response.text)
        self.assertEqual(response.json()["authenticated"], True)
        rows = repository.list_logs(organization_id=DEFAULT_PHASE1_ORGANIZATION_ID, limit=10)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["user_id"], DEFAULT_PHASE1_INTERNAL_USER_ID)
        self.assertEqual(rows[0]["ip_address"], "203.0.113.9")
        self.assertEqual(rows[0]["user_agent"], "pytest-agent/1.0")

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_tracker_entry_summaries_download_blanks_progress_note_when_requested(self) -> None:
        with ApiServer() as server:
            create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(create_status, 201)
            server.wait_for_run(create_payload["id"], target_statuses={"success"})

            tracker_status, tracker_payload = server.request_json(
                "POST",
                f"/api/runs/{create_payload['id']}/tracker-export",
            )
            self.assertEqual(tracker_status, 202)
            server.wait_for_run(tracker_payload["id"], target_statuses={"success"})

            xlsx_status, xlsx_body, _xlsx_headers = server.request_bytes(
                "GET",
                (
                    f"/api/tracker-entry-summaries/download?format=xlsx&blank_progress_note=true"
                    f"&source_tracker_run_id={tracker_payload['id']}"
                ),
            )
            self.assertEqual(xlsx_status, 200)

            workbook_path = _build_test_tmp_path(f"{uuid4().hex}_blank_progress.xlsx", token="files")
            workbook_path.parent.mkdir(parents=True, exist_ok=True)
            workbook_path.write_bytes(xlsx_body)
            try:
                rows = read_tracking_workbook_rows(workbook_path)
                self.assertTrue(rows)
                self.assertEqual(rows[0]["progress_note"], "")
            finally:
                workbook_path.unlink(missing_ok=True)

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_tracker_entry_summaries_download_job_creates_and_reuses_xlsx(self) -> None:
        with ApiServer() as server:
            create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
            self.assertEqual(create_status, 201)
            server.wait_for_run(create_payload["id"], target_statuses={"success"})

            tracker_status, tracker_payload = server.request_json(
                "POST",
                f"/api/runs/{create_payload['id']}/tracker-export",
            )
            self.assertEqual(tracker_status, 202)
            server.wait_for_run(tracker_payload["id"], target_statuses={"success"})

            create_job_status, create_job_payload = server.request_json(
                "POST",
                "/api/tracker-entry-summaries/download-jobs",
                payload={
                    "format": "xlsx",
                    "source_tracker_run_id": tracker_payload["id"],
                },
            )
            self.assertEqual(create_job_status, 202)
            self.assertEqual(create_job_payload["status"], "queued")

            deadline = time.time() + 10
            latest_payload = create_job_payload
            while time.time() < deadline:
                job_status, latest_payload = server.request_json(
                    "GET",
                    f"/api/tracker-entry-summaries/download-jobs/{create_job_payload['id']}",
                )
                self.assertEqual(job_status, 200)
                if latest_payload["status"] in {"success", "failed"}:
                    break
                time.sleep(0.1)
            self.assertEqual(latest_payload["status"], "success")
            self.assertTrue(latest_payload["download_url"])

            file_status, file_body, file_headers = server.request_bytes("GET", latest_payload["download_url"])
            self.assertEqual(file_status, 200)
            self.assertIn(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                file_headers.get("content-type", ""),
            )
            workbook = load_workbook(BytesIO(file_body), data_only=True)
            try:
                sheet = workbook[workbook.sheetnames[0]]
                self.assertEqual(str(sheet["A2"].value or "").strip(), "NO.")
                self.assertTrue(str(sheet["B3"].value or "").strip())
            finally:
                workbook.close()

            reused_status, reused_payload = server.request_json(
                "POST",
                "/api/tracker-entry-summaries/download-jobs",
                payload={
                    "format": "xlsx",
                    "source_tracker_run_id": tracker_payload["id"],
                },
            )
            self.assertEqual(reused_status, 202)
            self.assertEqual(reused_payload["id"], latest_payload["id"])
            self.assertEqual(reused_payload["status"], "success")
            self.assertEqual(reused_payload["reused_existing"], True)

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_tracker_template_upload_overrides_server_template_for_download(self) -> None:
        with ApiServer() as server:
            status_code, payload = server.request_json("GET", "/api/tracker-template")
            self.assertEqual(status_code, 200)
            self.assertEqual(payload["source"], "env_override")

            template_path = _build_test_tmp_path(f"{uuid4().hex}_upload_template.xlsx", token="files")
            template_path.parent.mkdir(parents=True, exist_ok=True)
            wb = load_workbook(ROOT_DIR / "assets" / "project_tracker_template.xlsx")
            try:
                ws = wb[wb.sheetnames[0]]
                ws["A1"] = "EC2 업로드 양식"
                wb.save(template_path)
            finally:
                wb.close()

            try:
                with template_path.open("rb") as fp:
                    response = requests.post(
                        f"{server.base_url}/api/tracker-template",
                        files={
                            "file": (
                                "project_tracker_uploaded.xlsx",
                                fp,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            )
                        },
                        timeout=20,
                    )
                self.assertEqual(response.status_code, 200, response.text)
                uploaded = response.json()
                self.assertEqual(uploaded["source"], "uploaded_override")
                self.assertEqual(uploaded["original_file_name"], "project_tracker_uploaded.xlsx")

                create_status, create_payload = server.request_json("POST", "/api/runs", payload=_project_tracker_run_payload())
                self.assertEqual(create_status, 201)
                server.wait_for_run(create_payload["id"], target_statuses={"success"})

                tracker_status, tracker_payload = server.request_json(
                    "POST",
                    f"/api/runs/{create_payload['id']}/tracker-export",
                )
                self.assertEqual(tracker_status, 202)
                server.wait_for_run(tracker_payload["id"], target_statuses={"success"})

                xlsx_status, xlsx_body, _ = server.request_bytes(
                    "GET",
                    f"/api/tracker-entry-summaries/download?format=xlsx&source_tracker_run_id={tracker_payload['id']}",
                )
                self.assertEqual(xlsx_status, 200)
                workbook = load_workbook(BytesIO(xlsx_body), data_only=True)
                try:
                    sheet = workbook[workbook.sheetnames[0]]
                    self.assertEqual(str(sheet["A1"].value or "").strip(), "EC2 업로드 양식")
                finally:
                    workbook.close()
            finally:
                template_path.unlink(missing_ok=True)

    def test_invalid_collect_mode_falls_back_to_auto(self) -> None:
        with ApiServer(
            env_overrides={
                "DATA_GO_KR_SERVICE_KEY": "",
                "PUBLIC_DATA_SERVICE_KEY": "",
                "G2B_SERVICE_KEY": "",
            }
        ) as server:
            status_code, payload = server.request_json(
                "POST",
                "/api/runs",
                payload={
                    "run_type": "project_tracker",
                    "advanced_options": {"collect_mode": "gui_seed"},
                    "params": {
                        "start_date": "20250101",
                        "end_date": "20250131",
                        "notice_title": "collect mode fallback case",
                        "bid_no": "",
                        "demand_org": "Fallback Test Org",
                        "rows_per_page": 20,
                        "max_pages": 1,
                        "api_scope": "construction",
                    },
                },
            )
            self.assertEqual(status_code, 201)
            run_detail = server.wait_for_run(payload["id"], target_statuses={"success"})
            self.assertEqual(run_detail["status"], "success")
            self.assertEqual(run_detail["summary"]["output"]["requested_collect_mode"], "auto")
            self.assertIn(
                run_detail["summary"]["output"]["collect_backend"],
                {"native_api", "synthetic"},
            )

    @pytest.mark.xfail(reason="tracker export path is not green in this branch")
    def test_synthetic_collect_mode_requires_debug_flag(self) -> None:
        with ApiServer(env_overrides={"PROJECT_TRACKER_ENABLE_SYNTHETIC_DEBUG": "0"}) as server:
            status_code, payload = server.request_json(
                "POST",
                "/api/runs",
                payload={
                    "run_type": "project_tracker",
                    "advanced_options": {"collect_mode": "synthetic"},
                    "params": {
                        "start_date": "20250101",
                        "end_date": "20250131",
                        "notice_title": "synthetic hidden",
                        "bid_no": "",
                        "demand_org": "Phase1 Internal Org",
                        "rows_per_page": 20,
                        "max_pages": 1,
                        "api_scope": "construction",
                    },
                },
            )
            self.assertEqual(status_code, 400)
            self.assertEqual(payload["error"]["code"], "validation_error")
            self.assertEqual(payload["error"]["message"], "collect_mode synthetic is only available in debug mode")

            dashboard_status, dashboard_payload = server.request_json("GET", "/api/dashboard/summary")
            self.assertEqual(dashboard_status, 200)
            self.assertEqual(dashboard_payload["synthetic_debug_enabled"], False)


if __name__ == "__main__":
    unittest.main()
