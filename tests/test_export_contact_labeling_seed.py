from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.export_contact_labeling_seed import _read_bucket_rows
from scripts.export_contact_labeling_seed import _write_xlsx


class ExportContactLabelingSeedTests(unittest.TestCase):
    def test_read_bucket_rows_seeds_label_columns(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            csv_path = Path(tmp_dir) / "hard.csv"
            with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=("project_name", "bid_no", "org", "contact"))
                writer.writeheader()
                writer.writerow(
                    {
                        "project_name": "예시 프로젝트",
                        "bid_no": "R25BK00000001",
                        "org": "예시군",
                        "contact": "건축과/055-000-0000",
                    }
                )

            rows = _read_bucket_rows(csv_path, "hard_wrong")
            self.assertEqual(len(rows), 1)
            self.assertEqual(rows[0]["seed_bucket"], "hard_wrong")
            self.assertEqual(rows[0]["candidate_text"], "건축과/055-000-0000")
            self.assertEqual(rows[0]["label_role"], "")
            self.assertEqual(rows[0]["label_status"], "")

    def test_write_xlsx_creates_bucket_sheets(self) -> None:
        rows = [
            {
                "seed_bucket": "hard_wrong",
                "seed_hint": "hint",
                "project_name": "예시 프로젝트",
                "bid_no": "R25BK00000001",
                "org": "예시군",
                "current_contact": "건축과/055-000-0000",
                "candidate_text": "건축과/055-000-0000",
                "label_role": "",
                "label_phase": "",
                "label_owner_side": "",
                "label_owner_side_basis": "",
                "label_final_pick_for_demand_contact": "",
                "label_status": "",
                "label_reason": "",
                "label_evidence_block": "",
                "reviewer_note": "",
            }
        ]

        with tempfile.TemporaryDirectory() as tmp_dir:
            xlsx_path = Path(tmp_dir) / "seed.xlsx"
            _write_xlsx(xlsx_path, rows)
            wb = load_workbook(xlsx_path)
            self.assertEqual(wb.sheetnames, ["전체", "하드 오입력", "약한 검토", "빈 연락처"])


if __name__ == "__main__":
    unittest.main()
