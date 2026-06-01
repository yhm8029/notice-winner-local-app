from __future__ import annotations

import csv
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from backend.services.native_tracker_backend import _derive_site_locations
from backend.services.native_tracker_backend import build_tracker_entries_from_winner_csv


class NativeTrackerBackendTests(unittest.TestCase):
    def test_derive_site_locations_uses_demand_org_name_over_project_garbage_tokens(self) -> None:
        region, city = _derive_site_locations(
            current_site_location="비엔날레전시",
            demand_org_name="광주광역시 동구",
            project_name="광주 비엔날레전시관 리모델링",
        )

        self.assertEqual(region, "광주광역시")
        self.assertEqual(city, "동구")

    def test_derive_site_locations_ignores_project_garbage_when_org_has_real_county(self) -> None:
        region, city = _derive_site_locations(
            current_site_location="합천운석충돌구",
            demand_org_name="경상남도 합천군",
            project_name="합천운석충돌구 미디어센터 건립",
        )

        self.assertEqual(region, "경상남도")
        self.assertEqual(city, "합천군")

    def test_derive_site_locations_uses_official_region_name_only(self) -> None:
        region, city = _derive_site_locations(
            current_site_location="재단법인경기도시장상권진흥원",
            demand_org_name="재단법인경기도시장상권진흥원",
            project_name="시장 활성화 사업",
        )

        self.assertEqual(region, "경기도")
        self.assertEqual(city, "")

    def test_derive_site_locations_filters_noise_and_keeps_real_location_examples(self) -> None:
        cases = [
            {
                "current_site_location": "",
                "demand_org_name": "경상북도",
                "project_name": "경상북도 사회복지회관 등 건립 설계공모",
                "expected": ("경상북도", ""),
            },
            {
                "current_site_location": "",
                "demand_org_name": "경상북도교육청",
                "project_name": "(집행대행)구펑초등학교 교사동 공간재구조화사업 설계공모",
                "expected": ("경상북도", ""),
            },
            {
                "current_site_location": "",
                "demand_org_name": "한국농어촌공사 전남지역본부 해남.완도지사",
                "project_name": "현산면 기초생활거점조성사업 세부설계용역 건축설계공모",
                "expected": ("전라남도", ""),
            },
            {
                "current_site_location": "",
                "demand_org_name": "전북특별자치도교육청 전북특별자치도정읍교육지원청",
                "project_name": "정읍제일고 학교복합문화센터 건립 설계공모",
                "expected": ("전북특별자치도", "정읍시"),
            },
            {
                "current_site_location": "",
                "demand_org_name": "대전광역시 건설관리본부",
                "project_name": "로봇드론지원센터 조성사업 기본 및 실시설계 용역 설계공모",
                "expected": ("대전광역시", ""),
            },
            {
                "current_site_location": "",
                "demand_org_name": "서울특별시교육청 서울특별시북부교육지원청",
                "project_name": "서울태릉초등학교 급식시설 및 기타 증축사업 설계용역[설계공모]",
                "expected": ("서울특별시", ""),
            },
            {
                "current_site_location": "",
                "demand_org_name": "서울특별시교육청 서울특별시강서양천교육지원청",
                "project_name": "경인초 급식실 및 학생식당, 기타시설 증축공사 설계공모",
                "expected": ("서울특별시", ""),
            },
            {
                "current_site_location": "",
                "demand_org_name": "충청남도교육청 충청남도논산계룡교육지원청",
                "project_name": "논산공업고등학교 그린스마트 미래학교 조성공사 제안 설계공모",
                "expected": ("충청남도", ""),
            },
            {
                "current_site_location": "",
                "demand_org_name": "당진도시공사",
                "project_name": "당진시 솔뫼공설묘지 묘역확장사업 기본 및 실시설계용역 설계공모 공고",
                "expected": ("", ""),
            },
            {
                "current_site_location": "",
                "demand_org_name": "경상북도 예천군 곤충연구소",
                "project_name": "예천군 곤충생태원 편의시설 조성사업 설계공모 공고",
                "expected": ("경상북도", "예천군"),
            },
            {
                "current_site_location": "",
                "demand_org_name": "울산광역시 울산경제자유구역청",
                "project_name": "경제자유구역 관련 사업",
                "expected": ("울산광역시", ""),
            },
            {
                "current_site_location": "",
                "demand_org_name": "문화재청 국립해양문화재연구소",
                "project_name": "연구소 시설 개선 사업",
                "expected": ("", ""),
            },
            {
                "current_site_location": "",
                "demand_org_name": "전북특별자치도 고창군",
                "project_name": "고창 동학농민혁명 성지화 사업 기본 및 실시설계 공모 제안서 제출 안내",
                "expected": ("전북특별자치도", "고창군"),
            },
            {
                "current_site_location": "",
                "demand_org_name": "대구광역시 동구",
                "project_name": "불로동 히트 조성사업 기본 및 실시설계용역 건축설계공모(간이공모)",
                "expected": ("대구광역시", "동구"),
            },
            {
                "current_site_location": "",
                "demand_org_name": "부산광역시해운대교육지원청",
                "project_name": "해운대여자중학교 공간재구조화 리모델링 건축설계 공모",
                "expected": ("부산광역시", "해운대구"),
            },
            {
                "current_site_location": "",
                "demand_org_name": "제주특별자치도 제주시",
                "project_name": "제주시 공공청사 건립 설계공모",
                "expected": ("제주특별자치도", "제주시"),
            },
            {
                "current_site_location": "",
                "demand_org_name": "경기도 수원시 장안구",
                "project_name": "조원1동 복합문화센터 건축 설계공모",
                "expected": ("경기도", "장안구"),
            },
        ]

        for case in cases:
            with self.subTest(project_name=case["project_name"]):
                region, city = _derive_site_locations(
                    current_site_location=case["current_site_location"],
                    demand_org_name=case["demand_org_name"],
                    project_name=case["project_name"],
                )
                self.assertEqual((region, city), case["expected"])

    def test_build_tracker_entries_from_winner_csv_maps_confirmed_fields(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "project_name_norm",
                        "post_title",
                        "winner_name",
                        "contract_name",
                        "notice_construction_cost",
                        "notice_construction_cost_source",
                        "contract_amount",
                        "contract_amount_source",
                        "contract_date",
                        "gross_area_scale",
                        "gross_area_scale_source",
                        "demand_contact",
                        "demand_contact_source",
                        "client_location",
                        "client_location_source",
                        "site_location",
                        "site_location_source",
                        "architect_office",
                        "architect_office_source",
                        "construction_start_date",
                        "construction_start_date_source",
                        "building_automation_estimated_amount",
                        "building_automation_estimated_amount_source",
                        "reason_code",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00554120",
                        "bid_ord": "000",
                        "project_name_norm": "design-project",
                        "post_title": "설계공모 결과",
                        "winner_name": "테스트건축사사무소",
                        "contract_name": "설계공모 결과",
                        "notice_construction_cost": "12,345,678원",
                        "notice_construction_cost_source": "confirmed_extracted",
                        "contract_amount": "1,000,000원",
                        "contract_amount_source": "confirmed_extracted",
                        "contract_date": "20250120",
                        "gross_area_scale": "12,345㎡",
                        "gross_area_scale_source": "confirmed_extracted",
                        "demand_contact": "시설과 / 김담당",
                        "demand_contact_source": "confirmed_extracted",
                        "client_location": "서울특별시 중구 세종대로 110",
                        "client_location_source": "confirmed_extracted",
                        "site_location": "서울특별시 중구 세종로",
                        "site_location_source": "confirmed_extracted",
                        "architect_office": "테스트건축사사무소",
                        "architect_office_source": "confirmed_extracted",
                        "construction_start_date": "2025-04-01",
                        "construction_start_date_source": "confirmed_extracted",
                        "building_automation_estimated_amount": "12,000,000원",
                        "building_automation_estimated_amount_source": "confirmed_extracted",
                        "reason_code": "NATIVE_WEB_MATCH",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date", "opening_scheduled_date", "g2b_verified"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00554120",
                        "bid_ord": "000",
                        "project_name": "설계공모 결과",
                        "org_name": "서울특별시교육청",
                        "announce_date": "20250102",
                        "opening_scheduled_date": "20250328",
                        "g2b_verified": "Y",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["architect_office"], "테스트건축사사무소")
        self.assertEqual(rows[0]["demand_org_name"], "서울특별시교육청")
        self.assertEqual(rows[0]["construction_cost"], "0.12억원")
        self.assertEqual(rows[0]["gross_area_scale"], "12,345㎡")
        self.assertEqual(rows[0]["demand_contact"], "시설과 / 김담당")
        self.assertEqual(rows[0]["client_location"], "서울특별시교육청")
        self.assertEqual(rows[0]["site_location_1"], "서울특별시")
        self.assertEqual(rows[0]["site_location_2"], "중구")
        self.assertEqual(rows[0]["opening_scheduled_date"], "20250328")
        self.assertEqual(rows[0]["construction_start_date"], "2025-04-01")
        self.assertEqual(rows[0]["manager_name"], "")
        self.assertEqual(rows[0]["building_automation_estimated_amount"], "12,000,000원")

    def test_build_tracker_entries_keeps_non_contact_fallback_out_of_core_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "project_name_norm",
                        "contract_name",
                        "notice_construction_cost",
                        "notice_construction_cost_source",
                        "contract_amount",
                        "contract_amount_source",
                        "demand_contact",
                        "demand_contact_source",
                        "architect_office",
                        "architect_office_source",
                        "building_automation_estimated_amount",
                        "building_automation_estimated_amount_source",
                        "reason_code",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00554120",
                        "bid_ord": "000",
                        "project_name_norm": "design-project",
                        "contract_name": "실제 프로젝트명",
                        "notice_construction_cost": "121.72억원",
                        "notice_construction_cost_source": "fallback_notice_cost",
                        "contract_amount": "497,745,454원",
                        "contract_amount_source": "fallback_seed_presmpt_prce",
                        "demand_contact": "박현진 / 055-960-2722",
                        "demand_contact_source": "fallback_seed_contact",
                        "architect_office": "테스트건축사사무소",
                        "architect_office_source": "fallback_winner_name",
                        "building_automation_estimated_amount": "497,745,454원",
                        "building_automation_estimated_amount_source": "fallback_seed_presmpt_prce",
                        "reason_code": "NATIVE_REVIEW_REQUIRED",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date", "spec_doc_file_name_1"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00554120",
                        "bid_ord": "000",
                        "project_name": "실제 프로젝트명",
                        "org_name": "경상남도교육청",
                        "announce_date": "20250102",
                        "spec_doc_file_name_1": "공고문.hwp",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["construction_cost"], "")
        self.assertEqual(rows[0]["demand_contact"], "박현진/055-960-2722")
        self.assertEqual(rows[0]["architect_office"], "")
        self.assertEqual(rows[0]["building_automation_estimated_amount"], "")
        self.assertIn("contract_amount<fallback_seed_presmpt_prce>=497,745,454원", rows[0]["progress_note"])
        self.assertIn("notice_construction_cost<fallback_notice_cost>=121.72억원", rows[0]["progress_note"])
        self.assertIn("demand_contact<fallback_seed_contact>=박현진 / 055-960-2722", rows[0]["progress_note"])
        self.assertIn("공고문.hwp", rows[0]["progress_note"])

    def test_build_tracker_entries_prefers_seed_project_name_over_generic_portal_title(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=["bid_no", "bid_ord", "contract_name", "post_title", "project_name_norm"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00554120",
                        "bid_ord": "000",
                        "contract_name": "나라장터",
                        "post_title": "접근 가능 브라우저 안내",
                        "project_name_norm": "portal-title",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00554120",
                        "bid_ord": "000",
                        "project_name": "경상남도교육청 함양도서관 이전 신축 설계공모",
                        "org_name": "경상남도교육청",
                        "announce_date": "20250102",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["project_name"], "경상남도교육청 함양도서관 이전 신축 설계공모")

    def test_build_tracker_entries_uses_estimated_building_automation_amount(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "notice_construction_cost",
                        "notice_construction_cost_source",
                        "building_automation_estimated_amount",
                        "building_automation_estimated_amount_source",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00554120",
                        "bid_ord": "000",
                        "contract_name": "실제 프로젝트명",
                        "notice_construction_cost": "12,172,404,000원",
                        "notice_construction_cost_source": "confirmed_extracted",
                        "building_automation_estimated_amount": "1.83억원~2.43억원",
                        "building_automation_estimated_amount_source": "estimated_notice_construction_cost",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00554120",
                        "bid_ord": "000",
                        "project_name": "실제 프로젝트명",
                        "org_name": "경상남도교육청 경상남도함양교육지원청",
                        "announce_date": "20250102",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["building_automation_estimated_amount"], "1.83억원~2.43억원")

    def test_build_tracker_entries_infers_client_location_from_org(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "client_location",
                        "client_location_source",
                        "site_location",
                        "site_location_source",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00570104",
                        "bid_ord": "000",
                        "contract_name": "실제 프로젝트명",
                        "client_location": "",
                        "client_location_source": "",
                        "site_location": "",
                        "site_location_source": "",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00570104",
                        "bid_ord": "000",
                        "project_name": "진주기계공업고등학교 실습장 및 청소년 직업체험관 증축 설계공모",
                        "org_name": "경상남도교육청 경상남도진주교육지원청",
                        "announce_date": "20250102",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["client_location"], "경상남도교육청 경상남도진주교육지원청")

    def test_build_tracker_entries_strips_parent_prefix_from_region_contact(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "demand_contact",
                        "demand_contact_source",
                        "site_location",
                        "site_location_source",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00569555",
                        "bid_ord": "000",
                        "contract_name": "양산소방서 119구조대 신축사업 설계공모",
                        "demand_contact": "소방본부 소방예산장비과 / 0552115694",
                        "demand_contact_source": "confirmed_extracted",
                        "site_location": "",
                        "site_location_source": "confirmed_extracted",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00569555",
                        "bid_ord": "000",
                        "project_name": "양산소방서 119구조대 신축사업 설계공모",
                        "org_name": "경상남도",
                        "announce_date": "20250115",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["demand_contact"], "소방예산장비과/055-211-5694")
        self.assertEqual(rows[0]["client_location"], "경상남도")
        self.assertEqual(rows[0]["site_location_1"], "경상남도")
        self.assertEqual(rows[0]["site_location_2"], "")

    def test_build_tracker_entries_blanks_person_only_contact(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=["bid_no", "bid_ord", "contract_name", "demand_contact", "demand_contact_source"],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00570104",
                        "bid_ord": "000",
                        "contract_name": "진주기계공업고등학교 실습장 및 청소년 직업체험관 증축 설계공모",
                        "demand_contact": "강경녀/055-740-2172",
                        "demand_contact_source": "confirmed_extracted",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00570104",
                        "bid_ord": "000",
                        "project_name": "진주기계공업고등학교 실습장 및 청소년 직업체험관 증축 설계공모",
                        "org_name": "경상남도교육청 경상남도진주교육지원청",
                        "announce_date": "20250110",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["demand_contact"], "")

    def test_build_tracker_entries_merges_existing_tracker_defaults(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"
            defaults_xlsx = root / "defaults.xlsx"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "gross_area_scale",
                        "gross_area_scale_source",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00569543",
                        "bid_ord": "000",
                        "contract_name": "함양소방서 안의119안전센터 신축사업 설계공모",
                        "gross_area_scale": "200㎡",
                        "gross_area_scale_source": "confirmed_extracted",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00569543",
                        "bid_ord": "000",
                        "project_name": "함양소방서 안의119안전센터 신축사업 설계공모",
                        "org_name": "경상남도",
                        "announce_date": "20250115",
                    }
                )

            wb = Workbook()
            ws = wb.active
            headers = [
                "NO.",
                "프로젝트명(시설비)",
                "연면적/규모",
                "공사비",
                "수요기관명",
                "수요기관(부서 및 담당자)",
                "발주처 위치",
                "현장 위치",
                "현장 위치",
                "설계사무소(건축)",
                "공사기간(착공일)",
                "최종 점검일자",
                "주요진행사항",
                "공고일",
                "담당자",
            ]
            for idx, value in enumerate(headers, start=1):
                ws.cell(2, idx).value = value
            ws.cell(3, 2).value = "함양소방서 안의119안전센터 신축사업 설계공모"
            ws.cell(3, 3).value = "119㎡"
            ws.cell(3, 7).value = "경상남도"
            ws.cell(3, 8).value = "경상남도"
            ws.cell(3, 9).value = "경상남도"
            wb.save(defaults_xlsx)

            with patch.dict(
                "os.environ",
                {
                    "TRACKER_MERGE_EXISTING_DEFAULTS": "1",
                    "TRACKER_DEFAULT_WORKBOOK_PATHS": str(defaults_xlsx),
                },
                clear=False,
            ):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["gross_area_scale"], "119㎡")
        self.assertEqual(rows[0]["client_location"], "경상남도")
        self.assertEqual(rows[0]["site_location_1"], "경상남도")
        self.assertEqual(rows[0]["site_location_2"], "")

    def test_build_tracker_entries_uses_trusted_contract_amount_when_notice_cost_is_blank(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "source_type",
                        "notice_construction_cost",
                        "notice_construction_cost_source",
                        "contract_amount",
                        "contract_amount_source",
                        "building_automation_estimated_amount",
                        "building_automation_estimated_amount_source",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00970325",
                        "bid_ord": "000",
                        "contract_name": "해운대여자중학교 공간재구조화 리모델링 건축설계 공모",
                        "source_type": "g2b_contract_api",
                        "notice_construction_cost": "",
                        "notice_construction_cost_source": "",
                        "contract_amount": "9,354,723,000원",
                        "contract_amount_source": "confirmed_contract_lookup",
                        "building_automation_estimated_amount": "",
                        "building_automation_estimated_amount_source": "",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00970325",
                        "bid_ord": "000",
                        "project_name": "해운대여자중학교 공간재구조화 리모델링 건축설계 공모",
                        "org_name": "부산광역시해운대교육지원청",
                        "announce_date": "20250721",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["construction_cost"], "93.55억원")
        self.assertEqual(rows[0]["building_automation_estimated_amount"], "1.40억원~1.87억원")

    def test_build_tracker_entries_persists_duration_and_computed_completion(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "source_type",
                        "contract_date",
                        "construction_start_date",
                        "construction_start_date_source",
                        "construction_duration_days",
                        "completion_expected_date_explicit",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00970325",
                        "bid_ord": "000",
                        "contract_name": "해운대여자중학교 공간재구조화 리모델링 건축설계 공모",
                        "source_type": "g2b_contract_api",
                        "contract_date": "2025-10-31",
                        "construction_start_date": "착수일로부터120일간",
                        "construction_start_date_source": "confirmed_extracted",
                        "construction_duration_days": "120",
                        "completion_expected_date_explicit": "",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00970325",
                        "bid_ord": "000",
                        "project_name": "해운대여자중학교 공간재구조화 리모델링 건축설계 공모",
                        "org_name": "부산광역시해운대교육지원청",
                        "announce_date": "20250721",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["contract_date"], "2025-10-31")
        self.assertEqual(rows[0]["construction_duration_days"], "120")
        self.assertEqual(rows[0]["completion_expected_date_explicit"], "")
        self.assertEqual(rows[0]["completion_expected_date_computed"], "2026-02-28")
        self.assertEqual(rows[0]["construction_start_date"], "계약일 2025-10-31 기준 120일 (완료예정 2026-02-28)")

    def test_build_tracker_entries_prefers_explicit_completion_over_computed_completion(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "source_type",
                        "contract_date",
                        "construction_start_date",
                        "construction_start_date_source",
                        "construction_duration_days",
                        "completion_expected_date_explicit",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00970325",
                        "bid_ord": "000",
                        "contract_name": "해운대여자중학교 공간재구조화 리모델링 건축설계 공모",
                        "source_type": "g2b_contract_api",
                        "contract_date": "2025-10-31",
                        "construction_start_date": "착수일로부터120일간",
                        "construction_start_date_source": "confirmed_extracted",
                        "construction_duration_days": "120",
                        "completion_expected_date_explicit": "2026-03-05",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00970325",
                        "bid_ord": "000",
                        "project_name": "해운대여자중학교 공간재구조화 리모델링 건축설계 공모",
                        "org_name": "부산광역시해운대교육지원청",
                        "announce_date": "20250721",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["completion_expected_date_explicit"], "2026-03-05")
        self.assertEqual(rows[0]["completion_expected_date_computed"], "")
        self.assertEqual(rows[0]["construction_start_date"], "계약일 2025-10-31 기준 120일 (완료예정 2026-03-05)")

    def test_build_tracker_entries_formats_notice_duration_without_contract_date(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "source_type",
                        "contract_date",
                        "construction_start_date",
                        "construction_start_date_source",
                        "construction_duration_days",
                        "completion_expected_date_explicit",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R26BK01311583",
                        "bid_ord": "000",
                        "contract_name": "duration-only design competition",
                        "source_type": "native_web",
                        "contract_date": "",
                        "construction_start_date": "",
                        "construction_start_date_source": "",
                        "construction_duration_days": "180",
                        "completion_expected_date_explicit": "",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R26BK01311583",
                        "bid_ord": "000",
                        "project_name": "duration-only design competition",
                        "org_name": "test org",
                        "announce_date": "20260203",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["construction_duration_days"], "180")
        self.assertEqual(rows[0]["construction_start_date"], "착수일로부터180일")

    def test_build_tracker_entries_prefers_long_duration_over_short_notice_window(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "source_type",
                        "contract_date",
                        "construction_start_date",
                        "construction_start_date_source",
                        "construction_duration_days",
                        "completion_expected_date_explicit",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R26BK01282184",
                        "bid_ord": "000",
                        "contract_name": "notice window before service duration",
                        "source_type": "native_web",
                        "contract_date": "",
                        "construction_start_date": "착수일로부터7일",
                        "construction_start_date_source": "confirmed_extracted",
                        "construction_duration_days": "300",
                        "completion_expected_date_explicit": "",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R26BK01282184",
                        "bid_ord": "000",
                        "project_name": "notice window before service duration",
                        "org_name": "test org",
                        "announce_date": "20260120",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["construction_duration_days"], "300")
        self.assertEqual(rows[0]["construction_start_date"], "착수일로부터300일")


    def test_build_tracker_entries_blanks_implausibly_small_gross_area_for_building_project(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "gross_area_scale",
                        "gross_area_scale_source",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "20240106221",
                        "bid_ord": "000",
                        "contract_name": "신항고등학교 교사 신축 설계공모",
                        "gross_area_scale": "1㎡",
                        "gross_area_scale_source": "confirmed_extracted",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "20240106221",
                        "bid_ord": "000",
                        "project_name": "신항고등학교 교사 신축 설계공모",
                        "org_name": "경상남도교육청",
                        "announce_date": "20240109",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["gross_area_scale"], "")

    def test_build_tracker_entries_blanks_implausibly_small_cost_when_area_is_large(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            winner_csv = root / "winner.csv"
            seed_csv = root / "seed.csv"

            with winner_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(
                    fp,
                    fieldnames=[
                        "bid_no",
                        "bid_ord",
                        "contract_name",
                        "gross_area_scale",
                        "gross_area_scale_source",
                        "notice_construction_cost",
                        "notice_construction_cost_source",
                        "building_automation_estimated_amount",
                        "building_automation_estimated_amount_source",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00820979",
                        "bid_ord": "000",
                        "contract_name": "서산시 시청사 건립사업 설계공모",
                        "gross_area_scale": "38,443.26㎡",
                        "gross_area_scale_source": "confirmed_extracted",
                        "notice_construction_cost": "1억원",
                        "notice_construction_cost_source": "confirmed_extracted",
                        "building_automation_estimated_amount": "",
                        "building_automation_estimated_amount_source": "",
                    }
                )

            with seed_csv.open("w", encoding="utf-8-sig", newline="") as fp:
                writer = csv.DictWriter(fp, fieldnames=["bid_no", "bid_ord", "project_name", "org_name", "announce_date"])
                writer.writeheader()
                writer.writerow(
                    {
                        "bid_no": "R25BK00820979",
                        "bid_ord": "000",
                        "project_name": "서산시 시청사 건립사업 설계공모",
                        "org_name": "충청남도 서산시",
                        "announce_date": "20251231",
                    }
                )

            with patch.dict("os.environ", {"TRACKER_MERGE_EXISTING_DEFAULTS": "0"}, clear=False):
                rows = build_tracker_entries_from_winner_csv(winner_csv_path=winner_csv, seed_csv_path=seed_csv)

        self.assertEqual(rows[0]["gross_area_scale"], "38,443.26㎡")
        self.assertEqual(rows[0]["construction_cost"], "")
        self.assertEqual(rows[0]["building_automation_estimated_amount"], "")
