from __future__ import annotations

import json
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from backend.services import artifact_files


def test_write_json_artifact_preserves_existing_file_when_serialization_fails(monkeypatch, tmp_path):
    artifacts_root = tmp_path / "artifacts"
    monkeypatch.setenv("ARTIFACTS_ROOT", str(artifacts_root))
    run_id = uuid4()
    written = artifact_files.write_json_artifact(
        run_id=run_id,
        file_name="related_notices.json",
        payload={"status": "previous"},
    )

    class Unserializable:
        pass

    try:
        artifact_files.write_json_artifact(
            run_id=run_id,
            file_name="related_notices.json",
            payload={"status": Unserializable()},
        )
    except TypeError:
        pass
    else:
        raise AssertionError("write_json_artifact should raise serialization errors")

    assert json.loads(written.absolute_path.read_text(encoding="utf-8")) == {"status": "previous"}
    assert not list(written.absolute_path.parent.glob(".related_notices.json.*.tmp"))


def test_tracking_download_workbook_uses_split_sheets_and_blanks_progress(monkeypatch, tmp_path) -> None:
    template_path = tmp_path / "template.xlsx"
    _write_template(template_path)
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template_path))

    payload = artifact_files.build_tracking_download_workbook_bytes(
        rows=[
            {
                "project_name": "서울 프로젝트",
                "gross_area_scale": "1,000㎡",
                "construction_cost": "10억원",
                "client_location": "서울특별시",
                "site_location_1": "서울특별시",
                "demand_org_name": "서울특별시",
                "progress_note": "Native web match",
                "notice_date": "20260102",
            },
            {
                "project_name": "서울교육청 프로젝트",
                "gross_area_scale": "2,000㎡",
                "construction_cost": "20억원",
                "client_location": "서울특별시교육청",
                "site_location_1": "서울특별시",
                "demand_org_name": "서울특별시교육청",
                "progress_note": "Native web match",
                "notice_date": "20260103",
            },
        ]
    )

    workbook_path = tmp_path / "download.xlsx"
    workbook_path.write_bytes(payload)
    wb = load_workbook(workbook_path, data_only=True)
    try:
        assert wb.sheetnames[:3] == ["전체", "서울", "서울교육청"]
        assert wb["전체"].cell(3, 2).value == "서울 프로젝트"
        assert wb["전체"].cell(4, 2).value == "서울교육청 프로젝트"
        assert wb["서울"].cell(3, 2).value == "서울 프로젝트"
        assert wb["서울교육청"].cell(3, 2).value == "서울교육청 프로젝트"
        assert wb["전체"].cell(3, 13).value in (None, "")
        assert wb["전체"].cell(4, 13).value in (None, "")
    finally:
        wb.close()


def test_tracking_download_workbook_limits_sheets_to_selected_regions(tmp_path) -> None:
    payload = artifact_files.build_tracking_download_workbook_bytes(
        selected_regions="부산",
        rows=[
            {
                "project_name": "서울 프로젝트",
                "client_location": "서울특별시",
                "site_location_1": "서울특별시",
                "demand_org_name": "서울특별시",
            },
            {
                "project_name": "부산 프로젝트",
                "client_location": "부산광역시",
                "site_location_1": "부산광역시",
                "demand_org_name": "부산광역시",
            },
            {
                "project_name": "부산교육청 프로젝트",
                "client_location": "부산광역시교육청",
                "site_location_1": "부산광역시",
                "demand_org_name": "부산광역시교육청",
            },
            {
                "project_name": "대구 프로젝트",
                "client_location": "대구광역시",
                "site_location_1": "대구광역시",
                "demand_org_name": "대구광역시",
            },
        ],
    )

    workbook_path = tmp_path / "busan-only.xlsx"
    workbook_path.write_bytes(payload)
    wb = load_workbook(workbook_path, data_only=True)
    try:
        assert wb.sheetnames == ["전체", "부산", "부산교육청"]
        assert [wb["전체"].cell(row, 2).value for row in range(3, wb["전체"].max_row + 1)] == [
            "부산 프로젝트",
            "부산교육청 프로젝트",
        ]
    finally:
        wb.close()


def test_tracking_download_workbook_pairs_selected_region_and_education_sheets(tmp_path) -> None:
    payload = artifact_files.build_tracking_download_workbook_bytes(
        selected_regions="부산,대구,인천",
        rows=[
            {
                "project_name": "부산 프로젝트",
                "client_location": "부산광역시",
                "site_location_1": "부산광역시",
                "demand_org_name": "부산광역시",
            },
            {
                "project_name": "부산교육청 프로젝트",
                "client_location": "부산광역시교육청",
                "site_location_1": "부산광역시",
                "demand_org_name": "부산광역시교육청",
            },
            {
                "project_name": "대구 프로젝트",
                "client_location": "대구광역시",
                "site_location_1": "대구광역시",
                "demand_org_name": "대구광역시",
            },
            {
                "project_name": "대구교육청 프로젝트",
                "client_location": "대구광역시교육청",
                "site_location_1": "대구광역시",
                "demand_org_name": "대구광역시교육청",
            },
            {
                "project_name": "인천 프로젝트",
                "client_location": "인천광역시",
                "site_location_1": "인천광역시",
                "demand_org_name": "인천광역시",
            },
            {
                "project_name": "인천교육청 프로젝트",
                "client_location": "인천광역시교육청",
                "site_location_1": "인천광역시",
                "demand_org_name": "인천광역시교육청",
            },
            {
                "project_name": "서울 프로젝트",
                "client_location": "서울특별시",
                "site_location_1": "서울특별시",
                "demand_org_name": "서울특별시",
            },
        ],
    )

    workbook_path = tmp_path / "multi-region.xlsx"
    workbook_path.write_bytes(payload)
    wb = load_workbook(workbook_path, data_only=True)
    try:
        assert wb.sheetnames == ["전체", "부산", "부산교육청", "대구", "대구교육청", "인천", "인천교육청"]
        assert "서울 프로젝트" not in [
            wb["전체"].cell(row, 2).value
            for row in range(3, wb["전체"].max_row + 1)
        ]
    finally:
        wb.close()


def _write_template(path: Path, *, site_headers: tuple[str, str] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "일반관리"
    ws["P1"] = "=TODAY()"
    if site_headers is None:
        site_headers = ("현장\n위치", "현장\n위치")

    headers = [
        "NO.",
        "프로젝트명(시설비)",
        "연면적/\n규모",
        "공사비",
        "수요기관명",
        "수요기관(부서 및 담당자)",
        "발주처\n위치",
        site_headers[0],
        site_headers[1],
        "설계사무소(건축)",
        "공사기간\n(착공일)",
        "최종\n점검일자",
        "주요진행사항",
        "공고일",
        "담당자",
        "빌딩자동제어 추정 금액",
    ]
    for column, value in enumerate(headers, start=1):
        ws.cell(2, column).value = value

    body_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    last_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    for row in (3, 4):
        for column in range(1, 17):
            cell = ws.cell(row, column)
            cell.fill = body_fill if column < 16 else last_fill

    wb.save(path)
    wb.close()


def _assert_auto_thin_horizontal_only_border(cell) -> None:
    assert cell.border.left is None or cell.border.left.style is None
    assert cell.border.right is None or cell.border.right.style is None
    assert cell.border.top.style == "thin"
    assert cell.border.bottom.style == "thin"
    assert cell.border.top.color.type == "auto"
    assert cell.border.top.color.auto is True
    assert cell.border.bottom.color.type == "auto"
    assert cell.border.bottom.color.auto is True


def test_resolve_tracker_template_path_prefers_primary_template(monkeypatch, tmp_path: Path):
    primary = tmp_path / "primary.xlsx"
    legacy = tmp_path / "legacy.xlsx"
    _write_template(primary)
    _write_template(legacy)

    monkeypatch.delenv("TRACKER_TEMPLATE_PATH", raising=False)
    monkeypatch.setattr(artifact_files, "DEFAULT_TRACKER_TEMPLATE_PATH", primary)
    monkeypatch.setattr(artifact_files, "LEGACY_TRACKER_TEMPLATE_PATH", legacy)

    assert artifact_files.resolve_tracker_template_path() == primary


def test_uploaded_tracker_template_override_has_highest_priority(monkeypatch, tmp_path: Path):
    primary = tmp_path / "primary.xlsx"
    legacy = tmp_path / "legacy.xlsx"
    env_template = tmp_path / "env.xlsx"
    uploaded = tmp_path / "uploaded.xlsx"
    uploaded_meta = tmp_path / "uploaded.json"
    for path in (primary, legacy, env_template):
        _write_template(path)

    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(env_template))
    monkeypatch.setattr(artifact_files, "DEFAULT_TRACKER_TEMPLATE_PATH", primary)
    monkeypatch.setattr(artifact_files, "LEGACY_TRACKER_TEMPLATE_PATH", legacy)
    monkeypatch.setattr(artifact_files, "UPLOADED_TRACKER_TEMPLATE_PATH", uploaded)
    monkeypatch.setattr(artifact_files, "UPLOADED_TRACKER_TEMPLATE_META_PATH", uploaded_meta)

    artifact_files.save_uploaded_tracker_template(
        payload=primary.read_bytes(),
        original_file_name="사용자.xlsx",
    )

    status = artifact_files.describe_active_tracker_template()
    assert artifact_files.resolve_tracker_template_path() == uploaded
    assert status["source"] == "uploaded_override"
    assert status["original_file_name"] == "사용자.xlsx"


def test_build_tracking_workbook_expands_template_rows_with_same_style(monkeypatch, tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_workbook_bytes(
        rows=[
            {"project_name": "프로젝트 1", "gross_area_scale": "100㎡"},
            {"project_name": "프로젝트 2", "gross_area_scale": "200㎡"},
            {"project_name": "프로젝트 3", "gross_area_scale": "300㎡"},
            {"project_name": "프로젝트 4", "gross_area_scale": "400㎡"},
        ]
    )

    output = tmp_path / "output.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)
    ws = wb[wb.sheetnames[0]]

    assert ws.max_row >= 6
    assert ws["B3"].value == "프로젝트 1"
    assert ws["B6"].value == "프로젝트 4"
    assert ws["A5"].style_id == ws["A3"].style_id
    assert ws["P6"].style_id == ws["P3"].style_id

    wb.close()


def test_build_tracking_download_workbook_splits_region_sheets(monkeypatch, tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_download_workbook_bytes(
        rows=[
            {"project_name": "서울 프로젝트", "site_location_1": "서울특별시"},
            {"project_name": "부산 프로젝트", "site_location_1": "부산광역시"},
            {"project_name": "대구 프로젝트", "site_location_1": "대구광역시"},
        ]
    )

    output = tmp_path / "download.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)

    assert wb.sheetnames == ["전체", "서울", "부산", "대구"]
    assert wb["서울"]["B3"].value == "서울 프로젝트"
    assert wb["부산"]["B3"].value == "부산 프로젝트"
    assert wb["대구"]["B3"].value == "대구 프로젝트"

    wb.close()


def test_build_tracking_download_workbook_splits_ordinary_and_education_office_sheets(
    monkeypatch, tmp_path: Path
):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_download_workbook_bytes(
        rows=[
            {
                "project_name": "서울시청 보건소",
                "demand_org_name": "서울특별시",
                "site_location_1": "서울특별시",
            },
            {
                "project_name": "의령 복합센터",
                "demand_org_name": "경상남도 의령군",
                "site_location_1": "경상남도",
            },
            {
                "project_name": "창원 농업센터",
                "demand_org_name": "경상남도 창원시 농업기술센터",
                "site_location_1": "경상남도",
            },
            {
                "project_name": "서울 북부 교육지원청 청사",
                "demand_org_name": "서울특별시교육청 서울특별시북부교육지원청",
                "client_location": "서울특별시교육청 서울특별시북부교육지원청",
                "site_location_1": "서울특별시",
            },
            {
                "project_name": "창녕 교육지원청 청사",
                "demand_org_name": "경상남도교육청 경상남도창녕교육지원청",
                "client_location": "경상남도교육청 경상남도창녕교육지원청",
                "site_location_1": "경상남도",
            },
        ]
    )

    output = tmp_path / "download_mixed.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)

    assert wb.sheetnames == ["전체", "서울", "경남", "서울교육청", "경남교육청"]
    assert [
        wb["전체"]["B3"].value,
        wb["전체"]["B4"].value,
        wb["전체"]["B5"].value,
        wb["전체"]["B6"].value,
        wb["전체"]["B7"].value,
    ] == [
        "서울시청 보건소",
        "의령 복합센터",
        "창원 농업센터",
        "서울 북부 교육지원청 청사",
        "창녕 교육지원청 청사",
    ]
    assert wb["서울"]["B3"].value == "서울시청 보건소"
    assert wb["서울"]["B4"].value in (None, "")
    assert wb["경남"]["B3"].value == "의령 복합센터"
    assert wb["경남"]["B4"].value == "창원 농업센터"
    assert wb["경남"]["B5"].value in (None, "")
    assert wb["서울교육청"]["B3"].value == "서울 북부 교육지원청 청사"
    assert wb["경남교육청"]["B3"].value == "창녕 교육지원청 청사"

    wb.close()


def test_header_columns_accepts_split_site_location_labels(monkeypatch, tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _write_template(template, site_headers=("현장위치(시도)", "현장위치(시군구)"))
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    workbook_path = artifact_files.resolve_tracker_template_path()
    wb = load_workbook(workbook_path)
    ws = wb[wb.sheetnames[0]]

    columns = artifact_files._header_columns(ws)  # type: ignore[attr-defined]

    assert len(columns["site_loc_region"]) == 1
    assert len(columns["site_loc_city"]) == 1

    wb.close()


def test_derive_tracking_education_office_sheet_name_collapses_support_offices_to_top_level_region() -> None:
    assert artifact_files._derive_tracking_education_office_sheet_name(  # type: ignore[attr-defined]
        {"demand_org_name": "경상남도교육청 경상남도창녕교육지원청"}
    ) == "경남교육청"
    assert artifact_files._derive_tracking_education_office_sheet_name(  # type: ignore[attr-defined]
        {"demand_org_name": "서울특별시강서양천교육지원청"}
    ) == "서울교육청"
    assert artifact_files._derive_tracking_education_office_sheet_name(  # type: ignore[attr-defined]
        {"demand_org_name": "경상남도 의령군"}
    ) == ""


def test_derive_tracking_education_office_sheet_name_falls_back_to_client_location() -> None:
    assert artifact_files._derive_tracking_education_office_sheet_name(  # type: ignore[attr-defined]
        {
            "demand_org_name": "",
            "client_location": "부산광역시해운대교육지원청",
        }
    ) == "부산교육청"


def test_build_tracking_download_workbook_keeps_total_sheet_when_no_classified_region_exists(
    monkeypatch, tmp_path: Path
):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_download_workbook_bytes(
        rows=[
            {
                "project_name": "분류불가 프로젝트",
                "demand_org_name": "기관명 미상",
                "client_location": "위치 확인 필요",
            }
        ]
    )

    output = tmp_path / "download_unclassified.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)

    assert wb.sheetnames == ["전체"]
    assert wb["전체"]["B3"].value == "분류불가 프로젝트"

    wb.close()


def test_build_tracking_download_workbook_applies_standard_download_formatting(monkeypatch, tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_download_workbook_bytes(
        rows=[
            {
                "project_name": "서울시청 보건소",
                "demand_org_name": "서울특별시",
                "site_location_1": "서울특별시",
                "notice_date": "20260418",
            },
            {
                "project_name": "창녕 교육지원청 청사",
                "demand_org_name": "경상남도교육청 경상남도창녕교육지원청",
                "client_location": "경상남도교육청 경상남도창녕교육지원청",
                "site_location_1": "경상남도",
                "notice_date": "20260501",
            },
        ]
    )

    output = tmp_path / "download_formatting.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)

    assert wb.sheetnames == ["전체", "서울", "경남교육청"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        assert ws.auto_filter.ref == "A2:R4"
        for column_letter in ("G", "L", "M", "O"):
            assert ws.column_dimensions[column_letter].hidden is True
        assert ws["Q2"].value == "설계사무소(전기)"
        assert ws["R2"].value == "설계사무소(기계)"
        assert ws["A2"].font.sz == 10
        assert ws["B3"].font.sz == 10

    assert wb["전체"]["N3"].value == "2026-04-18"
    assert wb["전체"]["N4"].value == "2026-05-01"
    assert wb["서울"]["N3"].value == "2026-04-18"
    assert wb["경남교육청"]["N3"].value == "2026-05-01"

    wb.close()


def test_build_tracking_workbook_fills_site_city_from_client_location_when_missing(monkeypatch, tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _write_template(template, site_headers=("현장위치(시도)", "현장위치(시군구)"))
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_workbook_bytes(
        rows=[
            {
                "project_name": "합성 Link 센터 및 에너지놀이터 조성 설계공모 공고",
                "client_location": "경상남도 창원시",
                "site_location_1": "경상남도",
                "site_location_2": "",
            }
        ]
    )

    output = tmp_path / "output.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)
    ws = wb[wb.sheetnames[0]]

    assert ws["H3"].value == "경상남도"
    assert ws["I3"].value == "창원시"

    wb.close()


def test_build_tracking_workbook_keeps_existing_site_city(monkeypatch, tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _write_template(template, site_headers=("현장위치(시도)", "현장위치(시군구)"))
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_workbook_bytes(
        rows=[
            {
                "project_name": "합성 Link 센터 및 에너지놀이터 조성 설계공모 공고",
                "client_location": "경상남도 창원시",
                "site_location_1": "경상남도",
                "site_location_2": "의창구",
            }
        ]
    )

    output = tmp_path / "output.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)
    ws = wb[wb.sheetnames[0]]

    assert ws["I3"].value == "의창구"

    wb.close()


def test_build_tracking_workbook_normalizes_education_office_site_city(monkeypatch, tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _write_template(template, site_headers=("\ud604\uc7a5\uc704\uce58(\uc2dc\ub3c4)", "\ud604\uc7a5\uc704\uce58(\uc2dc\uad70\uad6c)"))
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_workbook_bytes(
        rows=[
            {
                "project_name": "\ud569\ucc9c \uad50\uc721\uc9c0\uc6d0\uccad \uccad\uc0ac \ubcf4\uc218\uacf5\uc0ac",
                "client_location": "\uacbd\uc0c1\ub0a8\ub3c4 \ud569\ucc9c\uad50\uc721\uc9c0\uc6d0\uccad",
                "site_location_1": "\uacbd\uc0c1\ub0a8\ub3c4",
                "site_location_2": "\ud569\ucc9c\uad50\uc721\uc9c0\uc6d0\uccad",
            }
        ]
    )

    output = tmp_path / "output_education_office.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)
    ws = wb[wb.sheetnames[0]]

    assert ws["I3"].value == "\ud569\ucc9c\uad70"

    wb.close()


def test_build_tracking_workbook_prefers_most_specific_site_city_from_existing_address(monkeypatch, tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _write_template(template, site_headers=("\ud604\uc7a5\uc704\uce58(\uc2dc\ub3c4)", "\ud604\uc7a5\uc704\uce58(\uc2dc\uad70\uad6c)"))
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_workbook_bytes(
        rows=[
            {
                "project_name": "\ubd88\ub85c\ub3d9 \ud788\ud2b8 \uc870\uc131\uc0ac\uc5c5 \uae30\ubcf8 \ubc0f \uc2e4\uc2dc\uc124\uacc4\uc6a9\uc5ed \uac74\ucd95\uc124\uacc4\uacf5\ubaa8(\uac04\uc774\uacf5\ubaa8)",
                "client_location": "\ub300\uad6c\uad11\uc5ed\uc2dc \ub3d9\uad6c",
                "site_location_1": "\ub300\uad6c\uad11\uc5ed\uc2dc",
                "site_location_2": "\ub300\uad6c\uad11\uc5ed\uc2dc \ub3d9\uad6c",
            },
            {
                "project_name": "\uc870\uc6d01\ub3d9 \ubcf5\ud569\ubb38\ud654\uc13c\ud130 \uac74\ucd95 \uc124\uacc4\uacf5\ubaa8",
                "client_location": "\uacbd\uae30\ub3c4 \uc218\uc6d0\uc2dc \uc7a5\uc548\uad6c",
                "site_location_1": "\uacbd\uae30\ub3c4",
                "site_location_2": "\uacbd\uae30\ub3c4 \uc218\uc6d0\uc2dc \uc7a5\uc548\uad6c",
            },
        ]
    )

    output = tmp_path / "output_address_site_city.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)
    ws = wb[wb.sheetnames[0]]

    assert ws["I3"].value == "\ub3d9\uad6c"
    assert ws["I4"].value == "\uc7a5\uc548\uad6c"

    wb.close()


def test_build_tracking_workbook_uses_client_location_to_refine_less_specific_site_city(monkeypatch, tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _write_template(template, site_headers=("\ud604\uc7a5\uc704\uce58(\uc2dc\ub3c4)", "\ud604\uc7a5\uc704\uce58(\uc2dc\uad70\uad6c)"))
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_workbook_bytes(
        rows=[
            {
                "project_name": "\ubd88\ub85c\ub3d9 \ud788\ud2b8 \uc870\uc131\uc0ac\uc5c5 \uae30\ubcf8 \ubc0f \uc2e4\uc2dc\uc124\uacc4\uc6a9\uc5ed \uac74\ucd95\uc124\uacc4\uacf5\ubaa8(\uac04\uc774\uacf5\ubaa8)",
                "client_location": "\ub300\uad6c\uad11\uc5ed\uc2dc \ub3d9\uad6c",
                "site_location_1": "\ub300\uad6c\uad11\uc5ed\uc2dc",
                "site_location_2": "\ub300\uad6c\uad11\uc5ed\uc2dc",
            },
            {
                "project_name": "\uc870\uc6d01\ub3d9 \ubcf5\ud569\ubb38\ud654\uc13c\ud130 \uac74\ucd95 \uc124\uacc4\uacf5\ubaa8",
                "client_location": "\uacbd\uae30\ub3c4 \uc218\uc6d0\uc2dc \uc7a5\uc548\uad6c",
                "site_location_1": "\uacbd\uae30\ub3c4",
                "site_location_2": "\uc218\uc6d0\uc2dc",
            },
        ]
    )

    output = tmp_path / "output_refined_site_city.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)
    ws = wb[wb.sheetnames[0]]

    assert ws["I3"].value == "\ub3d9\uad6c"
    assert ws["I4"].value == "\uc7a5\uc548\uad6c"

    wb.close()
def test_build_tracking_workbook_does_not_extract_false_positive_city_tokens(monkeypatch, tmp_path: Path):
    template = tmp_path / "template.xlsx"
    _write_template(template, site_headers=("\ud604\uc7a5\uc704\uce58(\uc2dc\ub3c4)", "\ud604\uc7a5\uc704\uce58(\uc2dc\uad70\uad6c)"))
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_workbook_bytes(
        rows=[
            {
                "project_name": "\uacf5\uac04\uc7ac\uad6c \uac1c\uc120 \uc0ac\uc5c5",
                "client_location": "\uc804\ub77c\ub0a8\ub3c4 \uacf5\uac04\uc7ac\uad6c",
                "site_location_1": "\uc804\ub77c\ub0a8\ub3c4",
                "site_location_2": "\uc804\ub77c\ub0a8\ub3c4 \uacf5\uac04\uc7ac\uad6c",
            },
            {
                "project_name": "\uc2e0\ub3c4\uc2dc\uac1c\ubc1c\uc9c0\uad6c \uc870\uc131\uc0ac\uc5c5",
                "client_location": "\uacbd\uc0c1\ubd81\ub3c4 \uc2e0\ub3c4\uc2dc\uac1c\ubc1c\uc9c0\uad6c",
                "site_location_1": "\uacbd\uc0c1\ubd81\ub3c4",
                "site_location_2": "\uacbd\uc0c1\ubd81\ub3c4 \uc2e0\ub3c4\uc2dc\uac1c\ubc1c\uc9c0\uad6c",
            },
        ]
    )

    output = tmp_path / "output_false_positive_site_city.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)
    ws = wb[wb.sheetnames[0]]

    assert ws["I3"].value in (None, "")
    assert ws["I4"].value in (None, "")

    wb.close()


def test_build_tracking_download_workbook_applies_filters_to_row_two_on_all_sheets(
    monkeypatch, tmp_path: Path
):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_download_workbook_bytes(
        rows=[
            {
                "project_name": "Seoul Project",
                "site_location_1": "서울특별시",
                "notice_date": "20250804",
            },
            {
                "project_name": "Busan Project",
                "site_location_1": "부산광역시",
                "notice_date": "20250805",
            },
        ]
    )

    output = tmp_path / "download_filters.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        expected_ref = f"A2:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}"
        assert ws.auto_filter.ref == expected_ref

    wb.close()


def test_build_tracking_download_workbook_removes_vertical_borders_from_first_row(
    monkeypatch, tmp_path: Path
):
    template = tmp_path / "template.xlsx"
    _write_template(template)
    monkeypatch.setenv("TRACKER_TEMPLATE_PATH", str(template))

    payload = artifact_files.build_tracking_download_workbook_bytes(
        rows=[
            {
                "project_name": "Seoul Project",
                "site_location_1": "서울특별시",
                "notice_date": "20250804",
            }
        ]
    )

    output = tmp_path / "download_first_row_borders.xlsx"
    output.write_bytes(payload)
    wb = load_workbook(output)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for coord in ("A1", "P1"):
            _assert_auto_thin_horizontal_only_border(ws[coord])

    wb.close()
