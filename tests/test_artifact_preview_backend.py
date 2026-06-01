from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from backend.services.artifact_preview_backend import build_artifact_preview_payload


class BuildArtifactPreviewPayloadTests(unittest.TestCase):
    def test_builds_csv_preview_payload_with_normalized_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / "winner.csv"
            file_path.write_text(
                "name,amount,notes\nalpha,10,\nbeta,,kept\n",
                encoding="utf-8",
            )

            payload = build_artifact_preview_payload(
                artifact_type="winner_csv",
                file_path=file_path,
                limit=1,
            )

        self.assertEqual(payload["kind"], "table")
        self.assertEqual(payload["format"], "csv")
        self.assertEqual(payload["artifact_type"], "winner_csv")
        self.assertEqual(payload["headers"], ["name", "amount", "notes"])
        self.assertEqual(
            payload["rows"],
            [
                {
                    "name": "alpha",
                    "amount": "10",
                    "notes": "",
                }
            ],
        )
        self.assertEqual(payload["total_rows"], 2)

    def test_builds_execution_manifest_preview(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / "execution-manifest.json"
            file_path.write_text(
                json.dumps({"status": "success", "artifacts": ["winner_csv"]}),
                encoding="utf-8",
            )

            payload = build_artifact_preview_payload(
                artifact_type="execution_manifest",
                file_path=file_path,
                limit=5,
            )

        self.assertEqual(
            payload,
            {
                "kind": "json",
                "payload": {"status": "success", "artifacts": ["winner_csv"]},
            },
        )

    def test_builds_tracking_excel_preview_skipping_blank_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / "tracking.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Tracker"
            worksheet.cell(1, 1, "Tracking Export")
            worksheet.cell(1, 2, "Tracking Export")
            worksheet.cell(2, 1, "NO.")
            worksheet.cell(2, 2, "Project")
            worksheet.cell(3, 1, 1)
            worksheet.cell(3, 2, "Alpha")
            worksheet.cell(4, 1, "")
            worksheet.cell(4, 2, "   ")
            worksheet.cell(5, 1, 2)
            worksheet.cell(5, 2, "Beta")
            workbook.save(file_path)
            workbook.close()

            payload = build_artifact_preview_payload(
                artifact_type="tracking_excel",
                file_path=file_path,
                limit=1,
            )

        self.assertEqual(payload["kind"], "tracker_workbook")
        self.assertEqual(payload["format"], "xlsx")
        self.assertEqual(payload["artifact_type"], "tracking_excel")
        self.assertEqual(payload["sheet_name"], "Tracker")
        self.assertEqual(payload["title_row"][:2], ["Tracking Export", "Tracking Export"])
        self.assertEqual(payload["header_row"][:2], ["NO.", "Project"])
        self.assertEqual(payload["rows"], [["1", "Alpha"]])
        self.assertEqual(payload["total_rows"], 2)
        self.assertEqual(len(payload["column_widths"]), 2)

    def test_unsupported_artifact_invokes_callback_before_value_error_fallback(self) -> None:
        messages: list[str] = []

        def unsupported_preview(message: str) -> None:
            messages.append(message)

        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / "unsupported.bin"
            file_path.write_bytes(b"binary")

            with self.assertRaisesRegex(
                ValueError,
                "preview is not supported for artifact_type=unsupported",
            ):
                build_artifact_preview_payload(
                    artifact_type="unsupported",
                    file_path=file_path,
                    limit=5,
                    unsupported_preview_fn=unsupported_preview,
                )

        self.assertEqual(
            messages,
            ["preview is not supported for artifact_type=unsupported"],
        )

    def test_unsupported_artifact_raises_value_error_without_callback(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            file_path = Path(tmp_dir) / "unsupported.bin"
            file_path.write_bytes(b"binary")

            with self.assertRaisesRegex(
                ValueError,
                "preview is not supported for artifact_type=unsupported",
            ):
                build_artifact_preview_payload(
                    artifact_type="unsupported",
                    file_path=file_path,
                    limit=5,
                )
