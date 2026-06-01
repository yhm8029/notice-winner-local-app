from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from backend.services import artifact_files


def _write_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "일반관리"
    ws["P1"] = "=TODAY()"
    headers = [
        "NO.",
        "프로젝트명(시설비)",
        "연면적/\n규모",
        "공사비",
        "수요기관명",
        "수요기관(부서 및 담당자)",
        "발주처\n위치",
        "현장위치(시도)",
        "현장위치(시군구)",
        "설계사무소(건축)",
        "공사기간\n(착공일)",
        "최종\n점검일자",
        "주요진행사항",
        "공고일",
        "담당자",
        "빌딩자동제어 추정 금액",
    ]
    for column, value in enumerate(headers, start=1):
        ws.cell(2, column).value = value

    body_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    last_fill = PatternFill(fill_type="solid", fgColor="DDEBF7")
    for row in (3, 4):
        for column in range(1, 17):
            cell = ws.cell(row, column)
            cell.fill = body_fill if column < 16 else last_fill

    wb.save(path)
    wb.close()


class ArtifactFilesSiteLocationFromDemandOrgTests(unittest.TestCase):
    def test_build_tracking_workbook_normalizes_split_site_location_headers(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            template = root / "template.xlsx"
            output = root / "output_headers.xlsx"
            _write_template(template)

            with patch.dict("os.environ", {"TRACKER_TEMPLATE_PATH": str(template)}, clear=False):
                payload = artifact_files.build_tracking_workbook_bytes(
                    rows=[
                        {
                            "project_name": "sample",
                            "site_location_1": "경기도",
                            "site_location_2": "포천시",
                        }
                    ]
                )

            output.write_bytes(payload)
            wb = load_workbook(output)
            ws = wb[wb.sheetnames[0]]

            self.assertEqual(ws["H2"].value, "현장위치(시도)")
            self.assertEqual(ws["I2"].value, "현장위치(시군구)")

            wb.close()

    def test_build_tracking_workbook_recomputes_site_city_from_demand_org_name_when_existing_value_is_garbage(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            template = root / "template.xlsx"
            output = root / "output.xlsx"
            _write_template(template)

            with patch.dict("os.environ", {"TRACKER_TEMPLATE_PATH": str(template)}, clear=False):
                payload = artifact_files.build_tracking_workbook_bytes(
                    rows=[
                        {
                            "project_name": "광주 비엔날레전시관 리모델링",
                            "demand_org_name": "광주광역시 동구",
                            "client_location": "광주광역시 동구",
                            "site_location_1": "광주광역시",
                            "site_location_2": "비엔날레전시",
                        },
                        {
                            "project_name": "합천운석충돌구 미디어센터 건립",
                            "demand_org_name": "경상남도 합천군",
                            "client_location": "경상남도 합천군",
                            "site_location_1": "경상남도",
                            "site_location_2": "합천운석충돌구",
                        },
                    ]
                )

            output.write_bytes(payload)
            wb = load_workbook(output)
            ws = wb[wb.sheetnames[0]]

            self.assertEqual(ws["I3"].value, "동구")
            self.assertEqual(ws["I4"].value, "합천군")

            wb.close()

    def test_build_tracking_workbook_blanks_ambiguous_org_backed_site_city_without_trusted_site_source(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            template = root / "template.xlsx"
            output = root / "output_ambiguous.xlsx"
            _write_template(template)

            with patch.dict("os.environ", {"TRACKER_TEMPLATE_PATH": str(template)}, clear=False):
                payload = artifact_files.build_tracking_workbook_bytes(
                    rows=[
                        {
                            "project_name": "서울시 교육환경 개선",
                            "demand_org_name": "서울특별시교육청",
                            "client_location": "서울특별시교육청",
                            "site_location_1": "서울특별시",
                            "site_location_2": "서울특별시 중구 세종로",
                        },
                        {
                            "project_name": "정관도서관 리모델링",
                            "demand_org_name": "부산광역시교육청",
                            "client_location": "부산광역시교육청",
                            "site_location_1": "부산광역시",
                            "site_location_2": "부산광역시 기장군 정관읍",
                        },
                    ]
                )

            output.write_bytes(payload)
            wb = load_workbook(output)
            ws = wb[wb.sheetnames[0]]

            self.assertIn(ws["I3"].value, (None, ""))
            self.assertIn(ws["I4"].value, (None, ""))

            wb.close()
