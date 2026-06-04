from __future__ import annotations

import json
import threading
from datetime import datetime
from datetime import timezone
from pathlib import Path
from typing import Any
from uuid import UUID

from backend.api.support.run_support import _json_default
from backend.api.support.run_support import resolve_artifact_path


ROOT_DIR = Path(__file__).resolve().parents[3]
RELATED_NOTICE_TRACE_LOG_PATH = ROOT_DIR / "output" / "debug" / "related_notice_trace.jsonl"
RELATED_NOTICE_TRACE_LOG_LOCK = threading.Lock()


def _load_openpyxl_workbook_class():
    from openpyxl import Workbook

    return Workbook


def load_notice_seed_row_by_bid(*, bid_no: str, bid_ord: str = "000") -> dict[str, str] | None:
    from backend.services.notice_file_view_backend import load_notice_seed_row_by_bid as impl

    return impl(bid_no=bid_no, bid_ord=bid_ord)


def _load_notice_view_helpers(*, load_notice_seed_row_by_bid_fn=load_notice_seed_row_by_bid):
    from backend.services.notice_file_view_backend import build_desktop_notice_loading_html
    from backend.services.notice_file_view_backend import build_notice_file_fallback_html
    from backend.services.notice_file_view_backend import build_synap_viewer_embed_html
    from backend.services.notice_file_view_backend import download_notice_attachment
    from backend.services.notice_file_view_backend import infer_notice_attachment_suffix
    from backend.services.notice_file_view_backend import open_external_browser_url
    from backend.services.notice_file_view_backend import render_hwp_notice_html
    from backend.services.notice_file_view_backend import resolve_notice_viewer_url
    from backend.services.notice_file_view_backend import select_primary_notice_attachment
    from backend.services.notice_view_backend import build_notice_view_payload

    return {
        "build_notice_view_payload": build_notice_view_payload,
        "build_desktop_notice_loading_html": build_desktop_notice_loading_html,
        "build_notice_file_fallback_html": build_notice_file_fallback_html,
        "build_synap_viewer_embed_html": build_synap_viewer_embed_html,
        "download_notice_attachment": download_notice_attachment,
        "infer_notice_attachment_suffix": infer_notice_attachment_suffix,
        "load_notice_seed_row_by_bid": load_notice_seed_row_by_bid_fn,
        "open_external_browser_url": open_external_browser_url,
        "render_hwp_notice_html": render_hwp_notice_html,
        "resolve_notice_viewer_url": resolve_notice_viewer_url,
        "select_primary_notice_attachment": select_primary_notice_attachment,
    }


def _load_related_notice_precompute_helper():
    from backend.services.run_execution import queue_related_notice_precompute_for_run

    return queue_related_notice_precompute_for_run


def queue_related_notice_precompute_for_run(run_id: UUID, *, project_key: str = "") -> bool:
    return _load_related_notice_precompute_helper()(run_id, project_key=project_key)


def build_tracking_download_workbook_bytes(*, rows: list[dict[str, Any]]) -> bytes:
    from backend.services.artifact_files import build_tracking_download_workbook_bytes as impl

    return impl(rows=rows)


def _load_artifact_file_helpers(
    *,
    resolve_artifact_path_fn=resolve_artifact_path,
    build_tracking_download_workbook_bytes_fn=build_tracking_download_workbook_bytes,
):
    from backend.services.artifact_files import TRACKING_EXPORT_FIELDNAMES
    from backend.services.artifact_files import clear_uploaded_tracker_template
    from backend.services.artifact_files import describe_active_tracker_template
    from backend.services.artifact_files import save_uploaded_tracker_template

    return {
        "resolve_artifact_path": resolve_artifact_path_fn,
        "build_tracking_download_workbook_bytes": build_tracking_download_workbook_bytes_fn,
        "clear_uploaded_tracker_template": clear_uploaded_tracker_template,
        "describe_active_tracker_template": describe_active_tracker_template,
        "save_uploaded_tracker_template": save_uploaded_tracker_template,
        "tracking_export_fieldnames": TRACKING_EXPORT_FIELDNAMES,
    }


def _load_artifact_preview_helpers():
    from backend.services.artifact_preview_backend import build_artifact_preview_payload
    from backend.services.artifact_read_backend import build_artifact_item_payload
    from backend.services.artifact_read_backend import build_artifact_preview_payload_for_artifact_row

    return {
        "build_artifact_item_payload": build_artifact_item_payload,
        "build_artifact_preview_payload_for_artifact_row": build_artifact_preview_payload_for_artifact_row,
        "build_artifact_preview_payload": build_artifact_preview_payload,
    }


def _append_related_notice_trace(
    *,
    trace_id: str,
    event: str,
    project_id: UUID | None = None,
    project: dict[str, Any] | None = None,
    payload: dict[str, Any] | None = None,
) -> None:
    trace_payload: dict[str, Any] = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "trace_id": trace_id,
        "event": event,
    }
    if project_id is not None:
        trace_payload["project_id"] = str(project_id)
    if project is not None:
        trace_payload["project_name"] = str(project.get("project_name") or "")
        trace_payload["project_search_name"] = str(project.get("project_search_name") or "")
        trace_payload["issuer_name"] = str(project.get("issuer_name") or "")
        trace_payload["project_key"] = str(project.get("_project_match_key") or "")
        trace_payload["run_ids"] = list(project.get("source_json", {}).get("run_ids") or [])
    if payload:
        trace_payload.update(payload)

    trace_path = RELATED_NOTICE_TRACE_LOG_PATH
    trace_path.parent.mkdir(parents=True, exist_ok=True)
    with RELATED_NOTICE_TRACE_LOG_LOCK:
        with trace_path.open("a", encoding="utf-8") as fp:
            fp.write(json.dumps(trace_payload, ensure_ascii=False, default=_json_default) + "\n")
