from __future__ import annotations

import json
import logging
import os
import threading
import time
from datetime import datetime
from datetime import timezone
from pathlib import Path
from typing import Any
from uuid import UUID

from fastapi import FastAPI
from fastapi import Request
from fastapi import status
from fastapi.responses import JSONResponse
from fastapi.responses import RedirectResponse

from backend.api import auth_runtime as auth_runtime_module
from backend.api import schemas as schemas_module
from backend.api.routers import catalog as catalog_handlers
from backend.api.routers.catalog import delete_tracker_template_override
from backend.api.routers.catalog import get_tracker_template_status
from backend.api.routers.catalog import upload_tracker_template
from backend.api.routers import dashboard_handlers
from backend.api.routers.home_bootstrap_handlers import get_home_bootstrap
from backend.api.routers import project_aggregate_handlers
from backend.api.routers.registration import register_routes
from backend.api.routers.runs_write import create_tracker_export_run
from backend.api.routers import home_bootstrap_handlers
from backend.api.routers import tracker_export_handlers
from backend.api.routers import tracker_cache_handlers
from backend.api.routers import tracker_presentation_handlers
from backend.api.routers import tracker_read_handlers
from backend.api.routers import tracker_read_support
from backend.api.support import app_http_support
from backend.api.support import app_compat_support
from backend.api.support import app_pure_helpers
from backend.api.support import app_runtime_routes_support
from backend.api.support import app_runtime_support
from backend.api.support import auth_app_support
from backend.api.support import dashboard_app_support
from backend.api.support import project_aggregate_support
from backend.api.support import projects_related_notice_support
from backend.api.support import run_support
from backend.api.support import runtime_common
from backend.api.support import sales_support
from backend.api.support import tracker_support
from backend.api.support import tracker_read_support as tracker_read_support_module
from backend.perf_runtime import ensure_request_id
from backend.perf_runtime import HTTP_PERF_LOGGER
from backend.perf_runtime import measure_stage
from backend.perf_runtime import SLOW_HTTP_REQUEST_SEC
from backend.phase1_defaults import load_phase1_identity
from backend.repositories import RelatedNoticePublicationRepositoryConfigError
from backend.repositories import describe_repository_backends
from backend.repositories import get_home_bootstrap_snapshot_repository
from backend.repositories import get_related_notice_cache_repository
from backend.repositories import get_related_notice_publication_repository
from backend.repositories.tracker_entries import tracker_entry_matches_region
from backend.repositories.tracker_entries import tracker_entry_matches_title_visibility
from backend.services import related_notice_read_model_backend as _related_notice_read_model_backend
from backend.services.notice_file_view_backend import load_notice_seed_row_by_bid
from backend.services.project_dashboard_backend import build_project_aggregates as _build_project_aggregates_impl
from backend.services.project_dashboard_backend import coerce_uuid_or_none as _coerce_uuid_or_none_impl
from backend.services.project_dashboard_backend import derive_tracker_entry_bid_identity as _derive_tracker_entry_bid_identity_dashboard_impl
from backend.services.project_dashboard_backend import normalize_tracker_bid_ord as _normalize_tracker_bid_ord_dashboard_impl
from backend.services.related_notice_response_cache import clear_related_notice_response_cache as _clear_related_notice_response_cache
from backend.services.related_notice_response_cache import get_related_notice_response_cache as _get_related_notice_response_cache
from backend.services.related_notice_response_cache import set_related_notice_response_cache as _set_related_notice_response_cache
from backend.services.related_notice_refresh_worker import ensure_related_notice_refresh_worker_started
from backend.services.related_notice_refresh_worker import wake_related_notice_refresh_worker
from backend.services.report_job_store import get_report_job as get_stored_report_job
from backend.services.report_job_store import list_report_jobs as list_stored_report_jobs
from backend.services.report_job_store import store_report_job
from backend.services.report_job_store import to_report_job_item
from backend.services.run_execution import queue_related_notice_precompute_for_run
from backend.services.run_execution import safely_precompute_related_notices_for_run
from backend.services.run_preset_store import list_run_presets as list_stored_run_presets
from backend.services.run_preset_store import store_run_preset
from backend.services.run_preset_store import to_run_preset_item
from backend.services.tracker_download_job_store import find_reusable_tracker_download_job
from backend.services.tracker_download_job_store import get_tracker_download_job as get_stored_tracker_download_job
from backend.services.tracker_download_job_store import store_tracker_download_job
from backend.services.tracker_download_job_store import to_tracker_download_job_item
from backend.services.tracker_download_job_store import update_tracker_download_job

app = FastAPI(title="Project Tracker API", version="0.1.0")
FRONTEND_DIR = Path(__file__).resolve().parents[2] / "frontend"
APP_ROOT = FRONTEND_DIR.parents[0]
TRACKER_DOWNLOAD_JOB_ROOT = Path(
    str(os.getenv("TRACKER_DOWNLOAD_JOB_ROOT") or (APP_ROOT / ".tmp-tracker-download-jobs"))
).resolve()
RELATED_NOTICE_TRACE_LOG_PATH = APP_ROOT / "output" / "debug" / "related_notice_trace.jsonl"
BACKFILL_CONFLICT_RESOLUTIONS = tracker_support.BACKFILL_CONFLICT_RESOLUTIONS
PROJECT_NAMESPACE = tracker_read_support_module.PROJECT_NAMESPACE
_HOME_BOOTSTRAP_SNAPSHOT_VERSION = 3
_HOME_BOOTSTRAP_TRACKER_PAGE_SIZE = 20
HOME_BOOTSTRAP_PERF_LOGGER = logging.getLogger("perf.home_bootstrap")

_DASHBOARD_SUMMARY_CACHE_LOCK = threading.Lock()
_DASHBOARD_SUMMARY_CACHE: tuple[float, dict[str, Any]] | None = None
_DASHBOARD_SUMMARY_CACHE_TTL_SEC = 15.0
_ORGANIZATION_ADMIN_BOOTSTRAP_CACHE_LOCK = threading.Lock()
_ORGANIZATION_ADMIN_BOOTSTRAP_CACHE: dict[str, tuple[float, dict[str, Any]]] = {}
_ORGANIZATION_ADMIN_BOOTSTRAP_CACHE_TTL_SEC = 20.0
_SALES_CLAIM_SUMMARY_BY_USER_CACHE_LOCK = threading.Lock()
_SALES_CLAIM_SUMMARY_BY_USER_CACHE: dict[str, tuple[float, dict[str, Any]]] = {}
_SALES_CLAIM_SUMMARY_BY_USER_CACHE_TTL_SEC = 15.0
_TRACKER_GLOBAL_ROWS_CACHE_LOCK = threading.Lock()
_TRACKER_GLOBAL_ROWS_CACHE: tuple[float, list[dict[str, Any]]] | None = None
_TRACKER_GLOBAL_ROWS_CACHE_BUILD_EVENT: threading.Event | None = None
_TRACKER_GLOBAL_ROWS_CACHE_SERIAL = 0
TRACKER_GLOBAL_ROWS_CACHE_TTL_SEC = 300.0
TRACKER_GLOBAL_ROWS_CACHE_WAIT_TIMEOUT_SEC = 90.0
_TRACKER_EXPORT_WORKBOOK_CACHE_LOCK = threading.Lock()
_TRACKER_EXPORT_WORKBOOK_CACHE: dict[str, tuple[float, bytes]] = {}
_TRACKER_EXPORT_WORKBOOK_CACHE_BUILD_EVENTS: dict[str, threading.Event] = {}
_TRACKER_EXPORT_WORKBOOK_CACHE_SERIAL = 0
TRACKER_EXPORT_WORKBOOK_CACHE_TTL_SEC = 900.0
TRACKER_EXPORT_WORKBOOK_CACHE_WAIT_TIMEOUT_SEC = 90.0
TRACKER_EXPORT_WORKBOOK_CACHE_MAX_ENTRIES = 32
TRACKER_MEMORY_SOFT_CAP_RSS_BYTES = 900 * 1024 * 1024
TRACKER_MEMORY_SOFT_CAP_COOLDOWN_SEC = 300.0
_TRACKER_MEMORY_SOFT_CAP_LAST_CLEAR_MONOTONIC: float | None = None

AUTH_EXEMPT_API_PATHS = app_http_support.AUTH_EXEMPT_API_PATHS
ApiError = runtime_common.ApiError
auth_is_enabled = auth_runtime_module.auth_is_enabled
read_auth_context = auth_runtime_module.read_auth_context


def _load_openpyxl_workbook_class():
    from openpyxl import Workbook

    return Workbook


def _load_run_execution_helpers():
    from backend.services.run_execution import queue_tracker_export_run_for_parent
    from backend.services.run_execution import safely_execute_project_tracker

    return queue_tracker_export_run_for_parent, safely_execute_project_tracker


def _load_notice_view_helpers():
    from backend.services.notice_file_view_backend import build_desktop_notice_loading_html
    from backend.services.notice_file_view_backend import build_notice_file_fallback_html
    from backend.services.notice_file_view_backend import build_synap_viewer_embed_html
    from backend.services.notice_file_view_backend import download_notice_attachment
    from backend.services.notice_file_view_backend import infer_notice_attachment_suffix
    from backend.services.notice_file_view_backend import render_hwp_notice_html
    from backend.services.notice_file_view_backend import resolve_notice_viewer_url
    from backend.services.notice_file_view_backend import select_primary_notice_attachment
    from backend.services.notice_view_backend import build_notice_view_payload

    return {
        "build_desktop_notice_loading_html": build_desktop_notice_loading_html,
        "build_notice_file_fallback_html": build_notice_file_fallback_html,
        "build_synap_viewer_embed_html": build_synap_viewer_embed_html,
        "build_notice_view_payload": build_notice_view_payload,
        "download_notice_attachment": download_notice_attachment,
        "infer_notice_attachment_suffix": infer_notice_attachment_suffix,
        "load_notice_seed_row_by_bid": load_notice_seed_row_by_bid,
        "render_hwp_notice_html": render_hwp_notice_html,
        "resolve_notice_viewer_url": resolve_notice_viewer_url,
        "select_primary_notice_attachment": select_primary_notice_attachment,
    }


def _load_related_notice_precompute_helper():
    return queue_related_notice_precompute_for_run


def _load_artifact_file_helpers():
    from backend.services.artifact_files import apply_standard_download_workbook_formatting
    from backend.services.artifact_files import build_tracking_download_workbook_bytes
    from backend.services.artifact_files import clear_uploaded_tracker_template
    from backend.services.artifact_files import describe_active_tracker_template
    from backend.services.artifact_files import resolve_artifact_path
    from backend.services.artifact_files import save_uploaded_tracker_template
    from backend.services.artifact_files import TRACKING_EXPORT_FIELDNAMES

    return {
        "apply_standard_download_workbook_formatting": apply_standard_download_workbook_formatting,
        "build_tracking_download_workbook_bytes": build_tracking_download_workbook_bytes,
        "clear_uploaded_tracker_template": clear_uploaded_tracker_template,
        "describe_active_tracker_template": describe_active_tracker_template,
        "resolve_artifact_path": resolve_artifact_path,
        "save_uploaded_tracker_template": save_uploaded_tracker_template,
        "tracking_export_fieldnames": TRACKING_EXPORT_FIELDNAMES,
    }


def resolve_artifact_path(storage_path: str):
    return _load_artifact_file_helpers()["resolve_artifact_path"](storage_path)


def build_tracking_download_workbook_bytes(*, rows):
    return _load_artifact_file_helpers()["build_tracking_download_workbook_bytes"](rows=rows)


def _tracker_download_job_output_path(job_id: UUID) -> Path:
    TRACKER_DOWNLOAD_JOB_ROOT.mkdir(parents=True, exist_ok=True)
    return TRACKER_DOWNLOAD_JOB_ROOT / f"{job_id}.xlsx"


def _append_related_notice_trace(
    *,
    trace_id: str,
    event: str,
    project_id: UUID | None = None,
    project: dict[str, Any] | None = None,
    payload: dict[str, Any] | None = None,
) -> None:
    return app_runtime_support._append_related_notice_trace(
        trace_id=trace_id,
        event=event,
        project_id=project_id,
        project=project,
        payload=payload,
    )


def _auth_error_response(*, status_code: int, code: str, message: str) -> JSONResponse:
    return app_http_support._auth_error_response(status_code=status_code, code=code, message=message)


def _get_related_notice_publication_repository():
    try:
        return get_related_notice_publication_repository()
    except RelatedNoticePublicationRepositoryConfigError as exc:
        runtime_common._repository_error(str(exc))


@app.middleware("http")
async def add_request_id_and_log_slow_requests(request: Request, call_next):
    return await app_http_support.add_request_id_and_log_slow_requests(request, call_next)


@app.middleware("http")
async def disable_app_asset_cache(request: Request, call_next):
    return await app_http_support.disable_app_asset_cache(request, call_next)


@app.middleware("http")
async def enforce_phase2_auth(request: Request, call_next):
    return await app_http_support.enforce_phase2_auth(request, call_next)


@app.exception_handler(ApiError)
def handle_api_error(_request: Request, exc: ApiError) -> JSONResponse:
    return app_http_support.handle_api_error(_request, exc)


@app.exception_handler(Exception)
def handle_unhandled_error(_request: Request, exc: Exception) -> JSONResponse:
    return app_http_support.handle_unhandled_error(_request, exc)


@app.get("/health")
def healthcheck() -> dict[str, str]:
    return {"status": "ok"}


@app.on_event("startup")
def start_related_notice_refresh_worker() -> None:
    ensure_related_notice_refresh_worker_started(
        get_related_notice_cache_repository_fn=get_related_notice_cache_repository,
        safely_precompute_related_notices_for_run_fn=safely_precompute_related_notices_for_run,
    )


@app.get("/", include_in_schema=False)
def frontend_root() -> RedirectResponse:
    return RedirectResponse(url="/app/", status_code=status.HTTP_307_TEMPORARY_REDIRECT)


@app.get("/app", include_in_schema=False)
def frontend_app_root() -> RedirectResponse:
    return RedirectResponse(url="/app/", status_code=status.HTTP_307_TEMPORARY_REDIRECT)


_ATTRIBUTE_MODULES = (
    schemas_module,
    auth_runtime_module,
    runtime_common,
    app_runtime_routes_support,
    run_support,
    tracker_support,
    tracker_read_support_module,
    app_pure_helpers,
    app_compat_support,
    catalog_handlers,
    project_aggregate_handlers,
    project_aggregate_support,
    projects_related_notice_support,
    dashboard_handlers,
    dashboard_app_support,
    auth_app_support,
    sales_support,
    tracker_presentation_handlers,
    tracker_read_handlers,
    tracker_export_handlers,
    tracker_cache_handlers,
    home_bootstrap_handlers,
)


def __getattr__(name: str) -> Any:
    for module in _ATTRIBUTE_MODULES:
        if hasattr(module, name):
            return getattr(module, name)
    raise AttributeError(f"module {__name__!r} has no attribute {name!r}")


register_routes(app, frontend_dir=FRONTEND_DIR)
