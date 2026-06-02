from __future__ import annotations

import csv
import json
import os
import re
import shutil
from copy import copy
from io import BytesIO
from dataclasses import dataclass
from datetime import datetime
from datetime import timezone
from pathlib import Path
from typing import Any
from uuid import UUID
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.styles import Border
from openpyxl.styles import Color
from openpyxl.styles import Side
from openpyxl.utils import get_column_letter

from backend.repositories.tracker_entries import format_tracker_display_date
from backend.repositories.tracker_entries import TRACKER_REGION_ALIASES
from backend.repositories.tracker_entries import TRACKER_REGION_TOKEN_ONLY_CANONICALS
from backend.services import artifact_template_runtime
from backend.services.artifact_file_runtime import build_written_artifact
from backend.services.artifact_file_runtime import count_csv_rows
from backend.services.artifact_file_runtime import ensure_run_artifact_dir
from backend.services.artifact_file_runtime import resolve_artifact_path
from backend.services.artifact_workbook_formatting import apply_standard_download_workbook_formatting
from backend.services.native_gui_rules import OFFICIAL_REGION_PATTERN

ROOT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_TRACKER_TEMPLATE_PATH = ROOT_DIR / "프로젝트 트래커 양식.xlsx"
LEGACY_TRACKER_TEMPLATE_PATH = ROOT_DIR / "assets" / "project_tracker_template.xlsx"
UPLOADED_TRACKER_TEMPLATE_PATH = ROOT_DIR / "input" / "uploaded_project_tracker_template.xlsx"
UPLOADED_TRACKER_TEMPLATE_META_PATH = ROOT_DIR / "input" / "uploaded_project_tracker_template.json"
TRACKING_METRO_REGIONS = frozenset(
    {
        "서울특별시",
        "부산광역시",
        "대구광역시",
        "인천광역시",
        "광주광역시",
        "대전광역시",
        "울산광역시",
        "세종특별자치시",
        "제주특별자치도",
    }
)
TRACKING_INVALID_CITY_TOKEN_PARTS = (
    "지구",
    "재구",
    "구조",
    "개발",
    "조성",
    "사업",
    "센터",
    "학교",
    "청사",
    "시설",
    "공원",
    "단지",
    "권역",
    "문화",
)
DEFAULT_TRACKER_TEMPLATE_PATH = ROOT_DIR / "프로젝트 트랙커 양식.xlsx"
LEGACY_TRACKER_TEMPLATE_PATH = ROOT_DIR / "assets" / "project_tracker_template.xlsx"
UPLOADED_TRACKER_TEMPLATE_PATH = ROOT_DIR / "input" / "uploaded_project_tracker_template.xlsx"
UPLOADED_TRACKER_TEMPLATE_META_PATH = ROOT_DIR / "input" / "uploaded_project_tracker_template.json"

XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MIME_TYPE = "text/csv"
JSON_MIME_TYPE = "application/json"
TRACKING_EXPORT_FIELDNAMES = [
    "entry_key",
    "row_no",
    "source_bid_no",
    "source_bid_ord",
    "source_project_name_norm",
    "project_name",
    "gross_area_scale",
    "construction_cost",
    "demand_org_name",
    "demand_contact",
    "client_location",
    "site_location_1",
    "site_location_2",
    "architect_office",
    "construction_start_date",
    "last_checked_date",
    "progress_note",
    "notice_date",
    "manager_name",
    "building_automation_estimated_amount",
]
TRACKING_DOWNLOAD_EXTRA_HEADERS = (
    (17, "설계사무소(전기)"),
    (18, "설계사무소(기계)"),
)
TRACKING_DOWNLOAD_HIDDEN_COLUMNS = ("G", "L", "N")


@dataclass(frozen=True)
class WrittenArtifact:
    storage_path: str
    absolute_path: Path
    file_name: str
    mime_type: str
    size_bytes: int
    checksum: str
    row_count: int


def write_winner_csv(*, run_id: UUID, rows: list[dict[str, Any]]) -> WrittenArtifact:
    artifact_dir = ensure_run_artifact_dir(run_id)
    file_name = "project_tracker_rows.csv"
    absolute_path = artifact_dir / file_name
    with absolute_path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=TRACKING_EXPORT_FIELDNAMES)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: str(row.get(field, "")) for field in TRACKING_EXPORT_FIELDNAMES})

    return build_written_artifact(
        absolute_path=absolute_path,
        mime_type=CSV_MIME_TYPE,
        row_count=len(rows),
    )


def copy_csv_artifact(
    *,
    run_id: UUID,
    source_path: Path,
    artifact_file_name: str = "project_tracker_posts_files_v1_1.csv",
) -> WrittenArtifact:
    return copy_file_artifact(
        run_id=run_id,
        source_path=source_path,
        artifact_file_name=artifact_file_name,
        mime_type=CSV_MIME_TYPE,
        row_count=count_csv_rows(source_path),
    )


def copy_file_artifact(
    *,
    run_id: UUID,
    source_path: Path,
    artifact_file_name: str,
    mime_type: str,
    row_count: int = 0,
) -> WrittenArtifact:
    source_path = source_path.resolve()
    if not source_path.exists():
        raise FileNotFoundError(f"artifact source not found: {source_path}")

    artifact_dir = ensure_run_artifact_dir(run_id)
    absolute_path = artifact_dir / artifact_file_name
    shutil.copy2(source_path, absolute_path)
    return build_written_artifact(
        absolute_path=absolute_path,
        mime_type=mime_type,
        row_count=row_count,
    )


def write_json_artifact(
    *,
    run_id: UUID,
    file_name: str,
    payload: dict[str, Any] | list[Any],
) -> WrittenArtifact:
    artifact_dir = ensure_run_artifact_dir(run_id)
    absolute_path = artifact_dir / file_name
    temp_path = artifact_dir / f".{file_name}.{uuid4().hex}.tmp"
    try:
        with temp_path.open("w", encoding="utf-8") as fp:
            json.dump(payload, fp, ensure_ascii=False, indent=2)
            fp.flush()
            os.fsync(fp.fileno())
        os.replace(temp_path, absolute_path)
    finally:
        if temp_path.exists():
            temp_path.unlink()
    return build_written_artifact(
        absolute_path=absolute_path,
        mime_type=JSON_MIME_TYPE,
        row_count=0,
    )


def write_tracking_workbook(*, run_id: UUID, rows: list[dict[str, Any]]) -> WrittenArtifact:
    wb = _build_tracking_workbook(rows=rows)
    artifact_dir = ensure_run_artifact_dir(run_id)
    file_name = "project_tracking.xlsx"
    absolute_path = artifact_dir / file_name
    wb.save(absolute_path)

    return build_written_artifact(
        absolute_path=absolute_path,
        mime_type=XLSX_MIME_TYPE,
        row_count=len(rows),
    )


def build_tracking_workbook_bytes(*, rows: list[dict[str, Any]]) -> bytes:
    wb = _build_tracking_workbook(rows=rows)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_tracking_download_workbook_bytes(*, rows: list[dict[str, Any]]) -> bytes:
    wb = _build_tracking_workbook(rows=rows, split_region_sheets=True)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def read_tracking_workbook_rows(workbook_path: Path) -> list[dict[str, str]]:
    wb = load_workbook(workbook_path, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_cols = _header_columns(ws)
        rows: list[dict[str, str]] = []
        for row_idx in range(3, ws.max_row + 1):
            project_name = str(ws.cell(row_idx, header_cols["project_name"][0]).value or "").strip()
            if not project_name:
                continue
            rows.append(
                {
                    "project_name": project_name,
                    "gross_area_scale": _cell_text(ws, row_idx, header_cols["area"][0]),
                    "construction_cost": _cell_text(ws, row_idx, header_cols["cost"][0]),
                    "demand_org_name": _cell_text(ws, row_idx, header_cols["org_name"][0]),
                    "demand_contact": _cell_text(ws, row_idx, header_cols["contact"][0]),
                    "client_location": _cell_text(ws, row_idx, header_cols["issuer_loc"][0]),
                    "site_location_1": _cell_text(ws, row_idx, header_cols["site_loc_region"][0]),
                    "site_location_2": _cell_text(ws, row_idx, header_cols["site_loc_city"][0]),
                    "architect_office": _cell_text(ws, row_idx, header_cols["winner_name"][0]),
                    "construction_start_date": _cell_text(ws, row_idx, header_cols["construction_period"][0]),
                    "last_checked_date": _cell_text(ws, row_idx, header_cols["final_date"][0]),
                    "progress_note": _cell_text(ws, row_idx, header_cols["progress"][0]),
                    "notice_date": _cell_text(ws, row_idx, header_cols["announce_date"][0]),
                    "manager_name": _cell_text(ws, row_idx, header_cols["owner"][0]),
                    "building_automation_estimated_amount": (
                        _cell_text(ws, row_idx, header_cols["building_auto_est"][0])
                        if "building_auto_est" in header_cols
                        else ""
                    ),
                }
            )
        return rows
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _build_tracking_workbook(*, rows: list[dict[str, Any]], split_region_sheets: bool = False):
    template_path = resolve_tracker_template_path()
    wb = load_workbook(template_path)
    base_ws = wb[wb.sheetnames[0]]
    if not split_region_sheets:
        _populate_tracking_sheet(base_ws, rows=rows)
        apply_standard_download_workbook_formatting(wb)
        _apply_tracking_workbook_filters(wb)
        _apply_tracking_workbook_borders(wb)
        return wb

    download_template_ws = wb.copy_worksheet(base_ws)
    download_template_ws.title = "__download_template__"
    _populate_tracking_sheet(base_ws, rows=rows)
    used_titles: set[str] = set()
    base_ws.title = _make_tracking_sheet_title("전체", used_titles)
    ordinary_grouped, education_grouped = _group_tracking_rows_for_download_sheets(rows)
    if not ordinary_grouped and not education_grouped:
        wb.remove(download_template_ws)
        _apply_tracking_download_sheet_layout(base_ws)
        apply_standard_download_workbook_formatting(wb)
        _apply_tracking_workbook_filters(wb)
        _apply_tracking_workbook_borders(wb)
        return wb

    for region_name, region_rows in ordinary_grouped.items():
        ws = wb.copy_worksheet(download_template_ws)
        ws.title = _make_tracking_sheet_title(_short_tracking_region_name(region_name), used_titles)
        _populate_tracking_sheet(ws, rows=region_rows)
    for sheet_name, sheet_rows in education_grouped.items():
        ws = wb.copy_worksheet(download_template_ws)
        ws.title = _make_tracking_sheet_title(sheet_name, used_titles)
        _populate_tracking_sheet(ws, rows=sheet_rows)
    wb.remove(download_template_ws)
    for ws in wb.worksheets:
        _apply_tracking_download_sheet_layout(ws)
    apply_standard_download_workbook_formatting(wb)
    _apply_tracking_workbook_filters(wb)
    _apply_tracking_workbook_borders(wb)
    return wb


def _apply_tracking_workbook_filters(wb: Any) -> None:
    for ws in wb.worksheets:
        _apply_tracking_sheet_filter(ws)


def _apply_tracking_sheet_filter(ws: Any) -> None:
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{max(ws.max_row, 2)}"


def _apply_tracking_workbook_borders(wb: Any) -> None:
    for ws in wb.worksheets:
        _apply_tracking_sheet_borders(ws)


def _apply_tracking_sheet_borders(ws: Any) -> None:
    thin_auto_side = Side(style="thin", color=Color(auto=True))
    thin_auto_border = Border(
        left=thin_auto_side,
        right=thin_auto_side,
        top=thin_auto_side,
        bottom=thin_auto_side,
    )
    thin_auto_horizontal_border = Border(
        top=thin_auto_side,
        bottom=thin_auto_side,
    )
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            ws.cell(row, column).border = (
                thin_auto_horizontal_border if row == 1 else thin_auto_border
            )


def _populate_tracking_sheet(ws: Any, *, rows: list[dict[str, Any]]) -> None:
    header_cols = _header_columns(ws)
    _normalize_tracking_split_site_headers(ws, header_cols)
    _clear_tracker_rows(ws, header_cols)

    out_row = 3
    for seq, row in enumerate(rows, start=1):
        _ensure_tracker_row_capacity(ws, target_row=out_row)
        _set_row_value(ws, out_row, header_cols["no"], str(seq))
        _set_row_value(ws, out_row, header_cols["project_name"], str(row.get("project_name", "")))
        _set_row_value(ws, out_row, header_cols["area"], str(row.get("gross_area_scale", "")))
        _set_row_value(ws, out_row, header_cols["cost"], str(row.get("construction_cost", "")))
        _set_row_value(ws, out_row, header_cols["org_name"], str(row.get("demand_org_name", "")))
        _set_row_value(ws, out_row, header_cols["contact"], str(row.get("demand_contact", "")))
        _set_row_value(ws, out_row, header_cols["issuer_loc"], str(row.get("client_location", "")))
        _set_row_value(ws, out_row, header_cols["site_loc_region"], str(row.get("site_location_1", "")))
        _set_row_value(ws, out_row, header_cols["site_loc_city"], _resolve_tracking_site_city(row))
        _set_row_value(ws, out_row, header_cols["winner_name"], str(row.get("architect_office", "")))
        _set_row_value(ws, out_row, header_cols["construction_period"], str(row.get("construction_start_date", "")))
        _set_row_value(ws, out_row, header_cols["final_date"], str(row.get("last_checked_date", "")))
        _set_row_value(ws, out_row, header_cols["progress"], str(row.get("progress_note", "")))
        _set_row_value(
            ws,
            out_row,
            header_cols["announce_date"],
            format_tracker_display_date(row.get("notice_date", "")),
        )
        _set_row_value(ws, out_row, header_cols["owner"], str(row.get("manager_name", "")))
        if "building_auto_est" in header_cols:
            _set_row_value(
                ws,
                out_row,
                header_cols["building_auto_est"],
                str(row.get("building_automation_estimated_amount", "")),
            )
        out_row += 1


def _group_tracking_rows_by_region(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        region_name = _derive_tracking_region_name(row)
        if not region_name:
            continue
        grouped.setdefault(region_name, []).append(row)
    return _sort_tracking_grouped_rows_by_region(grouped)


def _group_tracking_rows_for_download_sheets(
    rows: list[dict[str, Any]],
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    ordinary_grouped: dict[str, list[dict[str, Any]]] = {}
    education_grouped: dict[str, list[dict[str, Any]]] = {}

    for row in rows:
        education_sheet_name = _derive_tracking_education_office_sheet_name(row)
        if education_sheet_name:
            education_grouped.setdefault(education_sheet_name, []).append(row)
            continue

        region_name = _derive_tracking_region_name(row)
        if not region_name:
            continue
        ordinary_grouped.setdefault(region_name, []).append(row)

    return (
        _sort_tracking_grouped_rows_by_region(ordinary_grouped),
        _sort_tracking_grouped_rows_by_region(
            education_grouped,
            region_name_resolver=_canonical_tracking_region_name_from_education_sheet_name,
        ),
    )


def _sort_tracking_grouped_rows_by_region(
    grouped: dict[str, list[dict[str, Any]]],
    *,
    region_name_resolver: Any | None = None,
) -> dict[str, list[dict[str, Any]]]:
    region_order = {name: index for index, name in enumerate(TRACKER_REGION_ALIASES.keys())}
    resolver = region_name_resolver or (lambda item: item)
    return {
        region_name: grouped[region_name]
        for region_name in sorted(
            grouped.keys(),
            key=lambda item: (region_order.get(resolver(item), 999), item),
        )
    }


def _derive_tracking_region_name(row: dict[str, Any]) -> str:
    for field_name in ("site_location_1", "site_location_2", "client_location", "demand_org_name"):
        text = str(row.get(field_name, "") or "").strip()
        if not text:
            continue
        for canonical, aliases in TRACKER_REGION_ALIASES.items():
            for alias in aliases:
                if _tracking_region_alias_matches(text=text, canonical=canonical, alias=alias):
                    return canonical
    return ""


def _tracking_region_alias_matches(*, text: str, canonical: str, alias: str) -> bool:
    if alias != canonical or canonical not in TRACKER_REGION_TOKEN_ONLY_CANONICALS:
        return alias in text
    return bool(re.search(rf"(?<![0-9A-Za-z가-힣]){re.escape(alias)}(?![0-9A-Za-z가-힣])", text))


TRACKING_REGION_SHORT_NAMES = {
    "서울특별시": "서울",
    "부산광역시": "부산",
    "대구광역시": "대구",
    "인천광역시": "인천",
    "광주광역시": "광주",
    "대전광역시": "대전",
    "울산광역시": "울산",
    "세종특별자치시": "세종",
    "제주특별자치도": "제주",
    "경기도": "경기",
    "강원특별자치도": "강원",
    "강원도": "강원",
    "충청북도": "충북",
    "충청남도": "충남",
    "전북특별자치도": "전북",
    "전라북도": "전북",
    "전라남도": "전남",
    "경상북도": "경북",
    "경상남도": "경남",
}


def _iter_tracking_issuer_texts(row: dict[str, Any]) -> list[str]:
    ordered = (
        str(row.get("demand_org_name") or "").strip(),
        str(row.get("client_location") or "").strip(),
    )
    return [value for value in ordered if value]


def _is_tracking_education_office_text(text: str) -> bool:
    normalized = str(text or "").strip()
    return "교육청" in normalized or "교육지원청" in normalized


def _short_tracking_region_name(region_name: str) -> str:
    normalized = str(region_name or "").strip()
    return TRACKING_REGION_SHORT_NAMES.get(normalized, normalized)


def _canonical_tracking_region_name_from_education_sheet_name(sheet_name: str) -> str:
    normalized = str(sheet_name or "").strip()
    if not normalized.endswith("교육청"):
        return normalized

    short_region = normalized.removesuffix("교육청")
    for canonical in TRACKER_REGION_ALIASES.keys():
        if _short_tracking_region_name(canonical) == short_region:
            return canonical
    return normalized


def _derive_tracking_education_office_sheet_name(row: dict[str, Any]) -> str:
    for text in _iter_tracking_issuer_texts(row):
        if not _is_tracking_education_office_text(text):
            continue
        for canonical, aliases in TRACKER_REGION_ALIASES.items():
            for alias in aliases:
                if _tracking_region_alias_matches(text=text, canonical=canonical, alias=alias):
                    short_region = _short_tracking_region_name(canonical)
                    return f"{short_region}교육청" if short_region else ""
    return ""


def _make_tracking_sheet_title(raw_title: str, used_titles: set[str]) -> str:
    title = re.sub(r"[:\\\\/?*\\[\\]]", "_", str(raw_title or "").strip())[:31] or "시트"
    candidate = title
    suffix = 2
    while candidate in used_titles:
        trimmed = title[: max(0, 31 - len(str(suffix)) - 1)]
        candidate = f"{trimmed}_{suffix}"
        suffix += 1
    used_titles.add(candidate)
    return candidate


def resolve_tracker_template_path() -> Path:
    _sync_artifact_template_runtime()
    return artifact_template_runtime.resolve_tracker_template_path()


def describe_active_tracker_template() -> dict[str, Any]:
    _sync_artifact_template_runtime()
    return artifact_template_runtime.describe_active_tracker_template()


def save_uploaded_tracker_template(*, payload: bytes, original_file_name: str = "") -> dict[str, Any]:
    if not payload:
        raise ValueError("template file is empty")

    lower_name = str(original_file_name or "").strip().lower()
    if lower_name and not lower_name.endswith(".xlsx"):
        raise ValueError("template file must be .xlsx")

    _validate_tracker_template_payload(payload)
    upload_path = resolve_uploaded_tracker_template_path()
    meta_path = resolve_uploaded_tracker_template_meta_path(upload_path)
    upload_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = upload_path.with_suffix(".tmp")
    temp_path.write_bytes(payload)
    temp_path.replace(upload_path)
    uploaded_at = datetime.now(timezone.utc).isoformat()
    meta = {
        "original_file_name": str(original_file_name or upload_path.name),
        "uploaded_at": uploaded_at,
        "size_bytes": len(payload),
    }
    meta_path.write_text(
        json.dumps(meta, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return describe_active_tracker_template()


def clear_uploaded_tracker_template() -> dict[str, Any]:
    _sync_artifact_template_runtime()
    return artifact_template_runtime.clear_uploaded_tracker_template()


def _cell_text(ws: Any, row_idx: int, col_idx: int) -> str:
    return str(ws.cell(row_idx, col_idx).value or "").strip()


def _resolve_tracking_site_city(row: dict[str, Any]) -> str:
    from backend.services.native_tracker_backend import normalize_tracker_site_locations

    _, site_city = normalize_tracker_site_locations(
        current_site_region=str(row.get("site_location_1", "") or "").strip(),
        current_site_city=str(row.get("site_location_2", "") or "").strip(),
        current_client_location=str(row.get("client_location", "") or "").strip(),
        demand_org_name=str(row.get("demand_org_name", "") or "").strip(),
        project_name=str(row.get("project_name", "") or "").strip(),
    )
    return site_city


def _split_tracking_region_city_from_address(value: str) -> tuple[str, str]:
    text = str(value or "").strip()
    if not text:
        return "", ""

    region_match = re.search(rf"({OFFICIAL_REGION_PATTERN})", text)
    region = str(region_match.group(1) or "").strip() if region_match else ""
    if not region_match:
        return "", ""

    suffix_text = text[region_match.end() :]
    tokens = re.findall(r"[가-힣]{1,20}", suffix_text)
    city = ""
    previous_city = ""
    for token in tokens:
        if _is_valid_tracking_city_token(token, region=region, previous_city=previous_city):
            city = token
            previous_city = token
    return region, city


def _is_valid_tracking_city_token(candidate: str, *, region: str, previous_city: str) -> bool:
    value = str(candidate or "").strip()
    if not value:
        return False
    if value.endswith(("특별시", "광역시", "특별자치시", "특별자치도")):
        return False
    if not re.fullmatch(r"[가-힣]{1,12}(?:시|군|구)", value):
        return False
    if any(token in value for token in TRACKING_INVALID_CITY_TOKEN_PARTS):
        return False
    if value.endswith("구"):
        return bool(previous_city.endswith("시") or region in TRACKING_METRO_REGIONS)
    return bool(region)


def _normalize_header(value: Any) -> str:
    return "".join(str(value or "").strip().split()).lower()


def _header_columns(ws: Any) -> dict[str, list[int]]:
    headers: dict[str, list[int]] = {}
    for column in range(1, ws.max_column + 1):
        normalized = _normalize_header(ws.cell(2, column).value)
        headers.setdefault(normalized, []).append(column)

    mapping = {
        "no": ("NO.",),
        "project_name": ("프로젝트명(시설비)",),
        "area": ("연면적/규모",),
        "cost": ("공사비", "예정공사비"),
        "org_name": ("수요기관명",),
        "contact": ("수요기관(부서및담당자)",),
        "issuer_loc": ("발주처위치",),
        "site_loc_region": ("현장위치(시도)", "현장위치"),
        "site_loc_city": ("현장위치(시군구)", "현장위치"),
        "winner_name": ("설계사무소(건축)",),
        "construction_period": ("공사기간(착공일)",),
        "final_date": ("최종입찰일자", "최종점검일자"),
        "progress": ("주요진행사항",),
        "announce_date": ("공고일",),
        "owner": ("담당자",),
        "building_auto_est": ("빌딩자동제어추정금액",),
    }

    resolved: dict[str, list[int]] = {}
    for logical_name, candidates in mapping.items():
        for candidate in candidates:
            cols = headers.get(_normalize_header(candidate))
            if cols:
                resolved[logical_name] = cols
                break

    required = (
        "no",
        "project_name",
        "area",
        "cost",
        "org_name",
        "contact",
        "issuer_loc",
        "site_loc_region",
        "site_loc_city",
        "winner_name",
        "construction_period",
        "final_date",
        "progress",
        "announce_date",
        "owner",
    )
    missing = [name for name in required if name not in resolved]
    if missing:
        raise RuntimeError(f"tracker template headers missing: {', '.join(missing)}")

    if len(resolved["site_loc_region"]) > 1:
        resolved["site_loc_region"] = [resolved["site_loc_region"][0]]
    if len(resolved["site_loc_city"]) > 1:
        resolved["site_loc_city"] = [resolved["site_loc_city"][1]]

    return resolved


def _normalize_tracking_split_site_headers(ws: Any, header_cols: dict[str, list[int]]) -> None:
    if header_cols.get("site_loc_region"):
        ws.cell(2, header_cols["site_loc_region"][0]).value = "현장위치(시도)"
    if header_cols.get("site_loc_city"):
        ws.cell(2, header_cols["site_loc_city"][0]).value = "현장위치(시군구)"


def _clear_tracker_rows(ws: Any, header_cols: dict[str, list[int]]) -> None:
    target_columns = sorted({column for columns in header_cols.values() for column in columns})
    for row in range(3, ws.max_row + 1):
        for column in target_columns:
            ws.cell(row, column).value = None


def _set_row_value(ws: Any, row: int, columns: list[int], value: str, index: int = 0) -> None:
    if not columns:
        return
    column_index = columns[min(index, len(columns) - 1)]
    ws.cell(row, column_index).value = value


def _apply_tracking_download_sheet_layout(ws: Any) -> None:
    _delete_tracking_progress_column(ws)
    for target_column, header in TRACKING_DOWNLOAD_EXTRA_HEADERS:
        adjusted_target_column = target_column - 1
        _copy_tracker_column_style(ws, source_column=15, target_column=adjusted_target_column)
        ws.cell(2, adjusted_target_column).value = header
        for row in range(3, ws.max_row + 1):
            ws.cell(row, adjusted_target_column).value = None
    for column_letter in TRACKING_DOWNLOAD_HIDDEN_COLUMNS:
        ws.column_dimensions[column_letter].hidden = True


def _delete_tracking_progress_column(ws: Any) -> None:
    header_cols = _header_columns(ws)
    progress_columns = header_cols.get("progress", [])
    if progress_columns:
        ws.delete_cols(progress_columns[0], 1)


def _copy_tracker_column_style(ws: Any, *, source_column: int, target_column: int) -> None:
    source_dimension = ws.column_dimensions[ws.cell(1, source_column).column_letter]
    target_dimension = ws.column_dimensions[ws.cell(1, target_column).column_letter]
    target_dimension.width = source_dimension.width
    target_dimension.hidden = False

    for row in range(1, ws.max_row + 1):
        source_cell = ws.cell(row, source_column)
        target_cell = ws.cell(row, target_column)
        target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)


def _ensure_tracker_row_capacity(ws: Any, *, target_row: int) -> None:
    if target_row <= ws.max_row:
        return

    template_row = 3 if ws.max_row >= 3 else ws.max_row
    for row in range(ws.max_row + 1, target_row + 1):
        _clone_tracker_template_row(ws, source_row=template_row, target_row=row)


def _clone_tracker_template_row(ws: Any, *, source_row: int, target_row: int) -> None:
    for column in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row, column)
        target_cell = ws.cell(target_row, column)
        target_cell._style = copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.font:
            target_cell.font = copy(source_cell.font)

    source_dimensions = ws.row_dimensions[source_row]
    target_dimensions = ws.row_dimensions[target_row]
    target_dimensions.height = source_dimensions.height
    target_dimensions.hidden = source_dimensions.hidden


def _validate_tracker_template_payload(payload: bytes) -> None:
    wb = load_workbook(BytesIO(payload))
    try:
        ws = wb[wb.sheetnames[0]]
        _header_columns(ws)
    except Exception as exc:
        raise ValueError(f"invalid tracker template: {exc}") from exc
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _tracker_template_source_label(source: str) -> str:
    mapping = {
        "uploaded_override": "업로드된 서버 양식",
        "env_override": "TRACKER_TEMPLATE_PATH 설정값",
        "repo_default": "repo 루트 기본 양식",
        "legacy_default": "legacy assets 양식",
    }
    return mapping.get(source, source)


def resolve_uploaded_tracker_template_path() -> Path:
    _sync_artifact_template_runtime()
    return artifact_template_runtime.resolve_uploaded_tracker_template_path()


def resolve_uploaded_tracker_template_meta_path(template_path: Path | None = None) -> Path:
    _sync_artifact_template_runtime()
    return artifact_template_runtime.resolve_uploaded_tracker_template_meta_path(template_path)


def _sync_artifact_template_runtime() -> None:
    artifact_template_runtime.DEFAULT_TRACKER_TEMPLATE_PATH = DEFAULT_TRACKER_TEMPLATE_PATH
    artifact_template_runtime.LEGACY_TRACKER_TEMPLATE_PATH = LEGACY_TRACKER_TEMPLATE_PATH
    artifact_template_runtime.UPLOADED_TRACKER_TEMPLATE_PATH = UPLOADED_TRACKER_TEMPLATE_PATH
    artifact_template_runtime.UPLOADED_TRACKER_TEMPLATE_META_PATH = UPLOADED_TRACKER_TEMPLATE_META_PATH
