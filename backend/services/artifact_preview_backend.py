from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def build_artifact_preview_payload(
    *,
    artifact_type: str,
    file_path: Path,
    limit: int,
    unsupported_preview_fn: Any | None = None,
) -> dict[str, Any]:
    if artifact_type == "execution_manifest":
        return {
            "kind": "json",
            "payload": json.loads(file_path.read_text(encoding="utf-8")),
        }
    if artifact_type in {"winner_csv", "candidate_csv", "internal_nav_csv", "seed_csv"}:
        return build_csv_preview_payload(file_path=file_path, limit=limit, artifact_type=artifact_type)
    if artifact_type == "tracking_excel":
        return build_tracking_excel_preview_payload(file_path=file_path, limit=limit)
    if unsupported_preview_fn is not None:
        unsupported_preview_fn(f"preview is not supported for artifact_type={artifact_type}")
    raise ValueError(f"preview is not supported for artifact_type={artifact_type}")


def build_csv_preview_payload(*, file_path: Path, limit: int, artifact_type: str) -> dict[str, Any]:
    with file_path.open("r", encoding="utf-8-sig", newline="") as fp:
        reader = csv.DictReader(fp)
        headers = list(reader.fieldnames or [])
        preview_rows: list[dict[str, str]] = []
        total_rows = 0
        for row in reader:
            total_rows += 1
            if len(preview_rows) < limit:
                preview_rows.append({header: str((row or {}).get(header) or "") for header in headers})
    return {
        "kind": "table",
        "format": "csv",
        "artifact_type": artifact_type,
        "headers": headers,
        "rows": preview_rows,
        "total_rows": total_rows,
    }


def build_tracking_excel_preview_payload(*, file_path: Path, limit: int) -> dict[str, Any]:
    wb = load_workbook(file_path, data_only=False)
    try:
        ws = wb[wb.sheetnames[0]]
        header_row = [str(ws.cell(2, col_idx).value or "") for col_idx in range(1, ws.max_column + 1)]
        title_row = [str(ws.cell(1, col_idx).value or "") for col_idx in range(1, ws.max_column + 1)]
        column_widths = [
            float(ws.column_dimensions[chr(64 + col_idx)].width or 12)
            for col_idx in range(1, min(ws.max_column, 26) + 1)
        ]

        preview_rows: list[list[str]] = []
        total_rows = 0
        for row_idx in range(3, ws.max_row + 1):
            row_values = [str(ws.cell(row_idx, col_idx).value or "") for col_idx in range(1, ws.max_column + 1)]
            if not any(value.strip() for value in row_values):
                continue
            total_rows += 1
            if len(preview_rows) < limit:
                preview_rows.append(row_values)

        return {
            "kind": "tracker_workbook",
            "format": "xlsx",
            "artifact_type": "tracking_excel",
            "sheet_name": ws.title,
            "title_row": title_row,
            "header_row": header_row,
            "rows": preview_rows,
            "total_rows": total_rows,
            "column_widths": column_widths,
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass
