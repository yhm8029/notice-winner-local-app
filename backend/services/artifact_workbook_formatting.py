from __future__ import annotations

import re
from copy import copy
from typing import Any

from openpyxl.utils import get_column_letter


def apply_standard_download_workbook_formatting(wb: Any) -> None:
    for ws in wb.worksheets:
        _apply_standard_download_sheet_formatting(ws)


def _apply_standard_download_sheet_formatting(ws: Any) -> None:
    if ws.max_column < 1 or ws.max_row < 2:
        return

    _set_download_sheet_font_size(ws, size=10)
    _normalize_download_sheet_notice_dates(ws)
    last_column = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A2:{last_column}{max(ws.max_row, 2)}"


def _set_download_sheet_font_size(ws: Any, *, size: int) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.value in (None, "") and not cell.has_style:
                continue
            font = copy(cell.font)
            font.sz = size
            cell.font = font


def _normalize_download_sheet_notice_dates(ws: Any) -> None:
    announce_columns = [
        column
        for column in range(1, ws.max_column + 1)
        if _normalize_header(ws.cell(2, column).value) == _normalize_header("공고일")
    ]
    if not announce_columns or ws.max_row < 3:
        return

    for column in announce_columns:
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row, column)
            cell.value = _format_download_notice_date(cell.value)


def _format_download_notice_date(value: Any) -> Any:
    text = str(value or "").strip()
    if not re.fullmatch(r"\d{8}", text):
        return value
    return f"{text[:4]}-{text[4:6]}-{text[6:]}"


def _normalize_header(value: Any) -> str:
    return "".join(str(value or "").strip().split()).lower()
